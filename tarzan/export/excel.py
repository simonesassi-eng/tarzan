"""Generate a professional multi-sheet Excel dashboard using openpyxl.

This module is the Reporting layer. It consumes a PortfolioMetrics object
and produces a formatted Excel workbook with charts, tables, and documentation.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tarzan.models.holding import Holding
from tarzan.models.investor_config import InvestorConfig
from tarzan.models.portfolio import PortfolioMetrics
from tarzan.export._format import ASSET_CLASS_COLORS, GEO_COLORS
from tarzan import config as cfg

logger = logging.getLogger(__name__)

SHEET_NAMES = cfg.sheet_names()


# ── STYLING ───────────────────────────────────────────

C = {
    'bg_page':   'F7F8FC',
    'bg_header': '1E293B',
    'bg_card':   'FFFFFF',
    'bg_alt':    'F8FAFF',
    'text_pri':  '1E293B',
    'text_sec':  '64748B',
    'text_wht':  'FFFFFF',
    'accent':    '5B5BD6',
    'green':     '16A34A',
    'red':       'DC2626',
    'amber':     'D97706',
    'border':    'CBD5E1',
    'border_dk': '94A3B8',
}

# Asset-class / geography colors come from the shared taxonomy in
# tarzan.export._format so the Excel dashboard and the HTML newsletter
# color the same class/region identically (single source of truth).
ASSET_COLORS = ASSET_CLASS_COLORS

TAB_COLORS = {
    'Dashboard': '5B5BD6', 'Holdings': '1E293B', 'Optimizer': '16A34A',
    'Performance': '2563EB', 'Return Contribution': 'D97706',
}

# KPI value color map
KPI_COLORS = {
    'Total Value': C['accent'], 'Sharpe': C['accent'], 'Sortino': C['accent'], 'Beta': C['accent'],
    'Total Gain': C['green'], 'CAGR': C['green'], 'RTD': C['green'], 'Alpha': C['green'],
    'Max Drawdown': C['red'], 'VaR': C['red'], 'CVaR': C['red'],
    'Volatility': C['amber'],
}


def px_fill(c):
    return PatternFill('solid', fgColor=c)


def px_font(size=10, bold=False, color='1E293B', italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)


def px_align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def px_border():
    s = Side(style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)


def px_border_hdr():
    s = Side(style='thin', color='94A3B8')
    return Border(left=s, right=s, top=s, bottom=s)


def px_no_border():
    return Border()


def _kpi_color(label):
    """Return the hex color for a KPI label based on KPI_COLORS map."""
    for key, color in KPI_COLORS.items():
        if key.lower() in label.lower():
            return color
    return C['accent']


def _num_color(val):
    """Return green or red hex color based on sign."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return C['text_pri']
    return C['green'] if val >= 0 else C['red']


def _deviation_color(delta_pct, tolerance):
    """Return a traffic-light color based on how far Actual% deviates from Target%.

    Args:
        delta_pct: Actual - Target (percentage points, signed).
        tolerance: Alert threshold in percentage points (from config.rebalancing_target_tolerance_pctg).

    Returns:
        Green if |delta| <= tolerance, amber if within 2× tolerance, red beyond.
        Neutral text color if delta or tolerance are unavailable.
    """
    if delta_pct is None or (isinstance(delta_pct, float) and pd.isna(delta_pct)):
        return C['text_pri']
    if tolerance is None or tolerance <= 0:
        return C['text_pri']
    abs_delta = abs(delta_pct)
    if abs_delta <= tolerance:
        return C['green']
    if abs_delta <= 2 * tolerance:
        return C['amber']
    return C['red']


def _apply_title(ws, row, col, text):
    """LIVELLO 1 — Sheet title."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = px_font(size=16, bold=True, color=C['text_pri'])
    cell.fill = px_fill(C['bg_page'])
    cell.alignment = px_align(h='left')
    cell.border = px_no_border()


def _apply_subtitle(ws, row, col, text):
    """LIVELLO 2 — Section subtitle."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = px_font(size=12, bold=True, color=C['text_pri'])
    cell.fill = px_fill(C['bg_page'])
    cell.alignment = px_align(h='left')
    cell.border = px_no_border()


def _apply_header(ws, row, col, text):
    """LIVELLO 3 — Table header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = px_font(size=10, bold=True, color=C['text_wht'])
    cell.fill = px_fill(C['bg_header'])
    cell.alignment = px_align(h='center')
    cell.border = px_border_hdr()


def _data_fill(table_idx):
    """Alternating row fill based on table-relative index (0-based)."""
    return px_fill(C['bg_card']) if table_idx % 2 == 0 else px_fill(C['bg_alt'])


def _write_data_cell(ws, row, col, value, table_idx, is_number=False, bold=False,
                     asset_class=None, geography=None, num_fmt=None, font_color=None):
    """Write a data cell with proper Clean Premium styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _data_fill(table_idx)
    cell.border = px_border()

    if font_color is not None:
        cell.font = px_font(size=10, bold=bold or is_number, color=font_color)
        cell.alignment = px_align(h='center' if is_number else 'left')
    elif asset_class and asset_class in ASSET_COLORS:
        cell.font = px_font(size=10, bold=True, color=ASSET_COLORS[asset_class])
        cell.alignment = px_align(h='left')
    elif geography and geography in GEO_COLORS:
        cell.font = px_font(size=10, bold=True, color=GEO_COLORS[geography])
        cell.alignment = px_align(h='left')
    elif is_number:
        color = _num_color(value)
        cell.font = px_font(size=10, bold=bold, color=color)
        cell.alignment = px_align(h='center')
    else:
        cell.font = px_font(size=10, bold=bold, color=C['text_pri'])
        cell.alignment = px_align(h='left')

    # Apply number format: explicit > auto-detect for numeric cells
    if num_fmt:
        cell.number_format = num_fmt
    elif is_number and isinstance(value, (int, float)):
        cell.number_format = '0.00'
    return cell


def _write_portfolio_row(ws, row, col_start, label, values, table_idx, num_fmt='0.00'):
    """Write a TOTAL PORTFOLIO / YOUR PORTFOLIO styled row."""
    portfolio_fill = px_fill('EEF2FF')
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.font = px_font(size=10, bold=True, color=C['accent'])
    cell.fill = portfolio_fill
    cell.border = px_border()
    cell.alignment = px_align(h='left')
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + 1 + i, value=val)
        c.fill = portfolio_fill
        c.border = px_border()
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            color = _num_color(val) if isinstance(val, (int, float)) else C['text_pri']
            c.font = px_font(size=10, bold=True, color=color)
            if isinstance(val, (int, float)) and num_fmt:
                c.number_format = num_fmt
        else:
            c.font = px_font(size=10, bold=True, color=C['text_pri'])
        c.alignment = px_align(h='center')


def _write_footer(ws, row, col):
    """Write the generation timestamp footer."""
    cell = ws.cell(row=row, column=col,
                   value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} v2.0")
    cell.font = px_font(size=8, italic=True, color=C['text_sec'])


def _format_number(val, is_pct=False):
    """Format a number for display, return string or None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "N/A"
    if is_pct:
        return round(val, 2)
    return round(val, 2) if isinstance(val, float) else val


# ── END STYLING ───────────────────────────────────────


def _make_bar(title, sheet, cat_col, series_defs, start_row, end_row, width=18, height=12):
    """Create a bulletproof bar chart. series_defs = [(col, label), ...]."""
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.width = width
    chart.height = height
    chart.style = 10

    cats = Reference(sheet, min_col=cat_col, min_row=start_row, max_row=end_row)
    for col, label in series_defs:
        vals = Reference(sheet, min_col=col, min_row=start_row, max_row=end_row)
        chart.add_data(vals, titles_from_data=False)
        chart.series[-1].tx = SeriesLabel(v=label)
    chart.set_categories(cats)

    return chart



def generate_excel(
    metrics: PortfolioMetrics,
    holdings: list[Holding],
    config: InvestorConfig,
    output_dir: str,
) -> str:
    """Generate the multi-sheet Excel dashboard.

    Args:
        metrics: Computed portfolio metrics.
        holdings: List of enriched holdings.
        config: Investor configuration.
        output_dir: Directory for the output file.

    Returns:
        Path to the generated Excel file.
    """
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    filepath = os.path.join(output_dir, f"portfolio_dashboard_{date_str}.xlsx")

    workbook = Workbook()
    try:
        # Create sheets (first sheet is created by default)
        sheets = {}
        for i, name in enumerate(SHEET_NAMES):
            if i == 0:
                ws = workbook.active
                ws.title = name
            else:
                ws = workbook.create_sheet(title=name)
            sheets[name] = ws
            # Global settings per sheet
            ws.sheet_view.showGridLines = False
            if name in TAB_COLORS:
                ws.sheet_properties.tabColor = TAB_COLORS[name]

        _write_dashboard(workbook, sheets["Dashboard"], metrics, config)
        _write_allocations(workbook, sheets["Optimizer"], metrics, config)
        _write_holdings(workbook, sheets["Holdings"], metrics)
        _write_performance(workbook, sheets["Performance"], metrics)
        _write_analysis(workbook, sheets["Return Contribution"], metrics)

        _set_column_widths(sheets)

        workbook.save(filepath)
        logger.info("Excel dashboard written to %s", filepath)
    except Exception as e:
        logger.error("Error generating Excel: %s", e)
        raise

    return filepath


def _set_column_widths(sheets: dict) -> None:
    """Set consistent column widths for all sheets."""
    s = sheets

    s["Dashboard"].column_dimensions['A'].width = 40
    s["Dashboard"].column_dimensions['B'].width = 16
    s["Dashboard"].column_dimensions['C'].width = 14
    s["Dashboard"].column_dimensions['D'].width = 3
    s["Dashboard"].column_dimensions['E'].width = 22
    s["Dashboard"].column_dimensions['F'].width = 16
    s["Dashboard"].column_dimensions['G'].width = 14
    for col_letter in ['H', 'I', 'J', 'K']:
        s["Dashboard"].column_dimensions[col_letter].width = 14

    s["Holdings"].column_dimensions['A'].width = 35
    s["Holdings"].column_dimensions['B'].width = 12
    s["Holdings"].column_dimensions['C'].width = 14
    s["Holdings"].column_dimensions['D'].width = 22
    s["Holdings"].column_dimensions['E'].width = 16
    s["Holdings"].column_dimensions['F'].width = 8
    for col_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        s["Holdings"].column_dimensions[col_letter].width = 12
    s["Holdings"].column_dimensions['O'].width = 10
    for col_letter in ['P', 'Q', 'R']:
        s["Holdings"].column_dimensions[col_letter].width = 16

    s["Optimizer"].column_dimensions['A'].width = 45
    s["Optimizer"].column_dimensions['B'].width = 12
    s["Optimizer"].column_dimensions['C'].width = 14
    s["Optimizer"].column_dimensions['D'].width = 12
    s["Optimizer"].column_dimensions['E'].width = 22

    s["Performance"].column_dimensions['A'].width = 35
    for i in range(2, 15):
        s["Performance"].column_dimensions[get_column_letter(i)].width = 10

    s["Return Contribution"].column_dimensions['A'].width = 35
    for i in range(2, 9):
        s["Return Contribution"].column_dimensions[get_column_letter(i)].width = 14


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------



def _write_dashboard(workbook, sheet, metrics: PortfolioMetrics, config: InvestorConfig):
    """Dashboard: Hero (Value/Gain/RTD) + Allocation + Top 5 + Rebalancing alert."""

    # --- Dates ---
    inception_str = ""
    if metrics.inception_date:
        try:
            inception_str = datetime.strptime(
                metrics.inception_date, "%Y-%m-%d"
            ).strftime("%d %b %Y")
        except ValueError:
            inception_str = metrics.inception_date
    as_of = datetime.now().strftime("%d %b %Y")

    # --- KPI values ---
    total_cost = float(metrics.holdings_df["cost_basis_eur"].sum()) if not metrics.holdings_df.empty else 0.0
    total_gain = metrics.total_value - total_cost
    rtd = (total_gain / total_cost * 100) if total_cost > 0 else 0.0

    # --- Column widths ---
    sheet.column_dimensions['A'].width = 22
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 14
    sheet.column_dimensions['D'].width = 3
    sheet.column_dimensions['E'].width = 22
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12

    # --- Title ---
    _apply_title(sheet, 1, 1, "Portfolio Dashboard")
    sub = sheet.cell(row=2, column=1, value=f"As of: {as_of}")
    sub.font = px_font(size=10, italic=True, color=C['text_sec'])
    sub.fill = px_fill(C['bg_page'])
    sub.border = px_no_border()

    # HERO - Portfolio Status
    row = 4
    inception_label = f" (since {inception_str})" if inception_str else ""
    _write_area_header(sheet, row, 1, 8, f"PORTFOLIO STATUS{inception_label}")
    row += 1

    hero_data = [
        ("Total Value (EUR)", metrics.total_value, None, "number"),
        ("Invested Value (EUR)", metrics.invested_value, None, "number"),
        ("Cash (EUR)", metrics.cash_value, None, "number"),
        ("Total Gain (EUR)", total_gain, total_gain, "number_signed"),
        ("RTD (%)", rtd, total_gain, "number_signed"),
    ]
    # Order-list returns: shown only when an order list was supplied
    # (xirr_pct is None otherwise, so a holdings-only run is unchanged).
    # Label the basis explicitly and keep the comparison like-for-like:
    # XIRR is annualized, so TWROR is shown both cumulative (since
    # inception) and annualized, so the reader compares annualized-with-
    # annualized rather than mistaking the cumulative TWROR for an
    # under-performance vs the annualized XIRR.
    if metrics.xirr_pct is not None:
        hero_data.append(
            ("XIRR / MWR (% ann.)", metrics.xirr_pct, metrics.xirr_pct, "pct_signed")
        )
    if metrics.twror_pct is not None:
        hero_data.append(
            ("TWROR (% cum. since inception)", metrics.twror_pct, metrics.twror_pct, "pct_signed")
        )
    if metrics.twror_annualized_pct is not None:
        hero_data.append(
            ("TWROR (% ann.)", metrics.twror_annualized_pct, metrics.twror_annualized_pct, "pct_signed")
        )
    # Net-of-tax ESTIMATE: shown only when a positive CGT was estimated.
    # Sits below the gross figures and is labeled "(est.)" so it is never
    # mistaken for the authoritative gross return.
    if metrics.estimated_cgt_eur is not None and metrics.estimated_cgt_eur > 0:
        hero_data.append(
            ("Est. CGT on realized gains (EUR)", -metrics.estimated_cgt_eur,
             -metrics.estimated_cgt_eur, "number_signed")
        )
        if metrics.pnl_eur_net_tax is not None:
            hero_data.append(
                ("PnL net of est. tax (EUR)", metrics.pnl_eur_net_tax,
                 metrics.pnl_eur_net_tax, "number_signed")
            )
        if metrics.xirr_net_tax_pct is not None:
            hero_data.append(
                ("XIRR net of est. tax (% ann.)", metrics.xirr_net_tax_pct,
                 metrics.xirr_net_tax_pct, "pct_signed")
            )
    for ti, (label, value, gain_for_color, kind) in enumerate(hero_data):
        lcell = sheet.cell(row=row, column=1, value=label)
        lcell.font = px_font(size=10, color=C['text_sec'])
        lcell.fill = _data_fill(ti)
        lcell.border = px_border()
        lcell.alignment = px_align(h='left')
        color = C['text_pri']
        if gain_for_color is not None and isinstance(gain_for_color, (int, float)):
            color = C['green'] if gain_for_color >= 0 else C['red']
        vcell = sheet.cell(row=row, column=2, value=value)
        vcell.font = px_font(size=11, bold=True, color=color)
        vcell.fill = _data_fill(ti)
        vcell.border = px_border()
        vcell.alignment = px_align(h='right')
        if kind == "number":
            vcell.number_format = '#,##0.00'
        elif kind == "number_signed":
            vcell.number_format = '+#,##0.00;-#,##0.00;0.00'
        elif kind == "pct_signed":
            vcell.number_format = '+0.00"%";-0.00"%";0.00"%"'
        row += 1

    # ALLOCATION
    row += 2
    _write_area_header(sheet, row, 1, 8, "ALLOCATION")
    row += 1

    ac_targets = {}
    geo_targets = {}
    if metrics.goal_deltas is not None and not metrics.goal_deltas.empty:
        for _, gd in metrics.goal_deltas.iterrows():
            if gd["type"] == "asset_class":
                ac_targets[gd["category"]] = (gd["target_pct"], gd["delta_pct"])
            elif gd["type"] == "geography (equity only)":
                geo_targets[gd["category"]] = (gd["target_pct"], gd["delta_pct"])
            # cash type is rendered in the Cash Buffer section, not here

    header_row = row
    for c, h in enumerate(["Asset Class", "Actual (% / EUR)", "Target (% / EUR)"], 1):
        _apply_header(sheet, header_row, c, h)
    row += 1

    tol = config.rebalancing_target_tolerance_pctg if config else 5.0
    if not metrics.allocation_by_class.empty:
        sorted_ac = metrics.allocation_by_class.sort_values("weight_pct", ascending=False)
        for ti, (_, rd) in enumerate(sorted_ac.iterrows()):
            cat = rd["category"]
            target_pct, delta_pct = ac_targets.get(cat, (None, None))
            dev_color = _deviation_color(delta_pct, tol)
            _write_data_cell(sheet, row, 1, cat, ti, asset_class=cat)
            _write_data_cell(sheet, row, 2, rd["weight_pct"], ti, is_number=True,
                             font_color=dev_color)
            _write_data_cell(sheet, row, 3,
                             target_pct if target_pct is not None else "",
                             ti, is_number=target_pct is not None,
                             font_color=C['text_pri'] if target_pct is not None else None)
            row += 1

    # Cash buffer row appended at the bottom of the Asset Class block
    # (hybrid EUR values sharing the same two columns).
    if config and config.target_cash_buffer_eur > 0:
        ti_cash = (len(metrics.allocation_by_class)
                   if not metrics.allocation_by_class.empty else 0)
        cash_tgt = float(config.target_cash_buffer_eur)
        cash_actual = metrics.cash_value
        cash_delta_rel_pctg = (
            (cash_actual - cash_tgt) / cash_tgt * 100.0 if cash_tgt > 0 else 0.0
        )
        cash_dev_color = _deviation_color(cash_delta_rel_pctg, tol)
        _write_data_cell(sheet, row, 1, "Cash & Cash Equivalents", ti_cash,
                         asset_class="Cash & Cash Equivalents")
        _write_data_cell(sheet, row, 2, cash_actual, ti_cash, is_number=True,
                         num_fmt='"€"#,##0.00', font_color=cash_dev_color)
        _write_data_cell(sheet, row, 3, cash_tgt, ti_cash, is_number=True,
                         num_fmt='"€"#,##0.00', font_color=C['text_pri'])
        row += 1

    ac_end_row = row

    row = header_row
    for c, h in enumerate(["Geography (Equity)", "Actual %", "Target %"], 5):
        _apply_header(sheet, row, c, h)
    row += 1

    if not metrics.allocation_by_geo.empty:
        sorted_geo = metrics.allocation_by_geo.sort_values("weight_pct", ascending=False)
        for ti, (_, rd) in enumerate(sorted_geo.iterrows()):
            cat = rd["category"]
            target_pct, delta_pct = geo_targets.get(cat, (None, None))
            dev_color = _deviation_color(delta_pct, tol)
            _write_data_cell(sheet, row, 5, cat, ti, geography=cat)
            _write_data_cell(sheet, row, 6, rd["weight_pct"], ti, is_number=True,
                             font_color=dev_color)
            _write_data_cell(sheet, row, 7,
                             target_pct if target_pct is not None else "",
                             ti, is_number=target_pct is not None,
                             font_color=C['text_pri'] if target_pct is not None else None)
            row += 1

    row = max(row, ac_end_row)

    # TOP 5 HOLDINGS
    row += 2
    _write_area_header(sheet, row, 1, 8, "TOP 5 HOLDINGS")
    row += 1

    for c, h in enumerate(["Name", "Value \u20ac", "Weight %", "Gain %", "Class"], 1):
        _apply_header(sheet, row, c, h)
    row += 1

    if not metrics.holdings_df.empty:
        top5 = metrics.holdings_df.nlargest(5, "weight_pct")
        for ti, (_, hr) in enumerate(top5.iterrows()):
            _write_data_cell(sheet, row, 1, hr.get("name", ""), ti)
            _write_data_cell(sheet, row, 2, hr.get("current_value", 0), ti, is_number=True, num_fmt='#,##0.00')
            _write_data_cell(sheet, row, 3, hr.get("weight_pct", 0), ti, is_number=True)
            _write_data_cell(sheet, row, 4, hr.get("gain_pct", 0), ti, is_number=True)
            _write_data_cell(sheet, row, 5, hr.get("asset_class", ""), ti, asset_class=hr.get("asset_class"))
            row += 1

    # REBALANCING ALERT
    if metrics.rebalancing_suggestions:
        row += 2
        n = len(metrics.rebalancing_suggestions)
        s_str = "s" if n > 1 else ""
        alert = f"\u26a0  {n} rebalancing action{s_str} suggested. See the Optimizer tab for details."
        cell = sheet.cell(row=row, column=1, value=alert)
        cell.font = px_font(size=11, bold=True, color=C['amber'])
        cell.fill = px_fill(C['bg_page'])
        cell.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    # Clear column D (spacer) for all dashboard rows
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import PatternFill
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for clear_r in range(1, row + 1):
        dc = sheet.cell(row=clear_r, column=4)
        if not isinstance(dc, MergedCell):
            dc.fill = white_fill
            dc.border = px_no_border()
            dc.value = None

    _write_footer(sheet, row + 1, 1)


def _write_area_header(ws, row, col_start, col_end, title):
    """Write a section header — clean subtitle style, consistent across all tabs."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font = px_font(size=12, bold=True, color=C['text_pri'])
    cell.fill = px_fill(C['bg_page'])
    cell.alignment = px_align(h='left', v='center')
    cell.border = px_no_border()
    # Bottom border accent
    from openpyxl.styles import Border, Side
    accent_border = Border(bottom=Side(style='medium', color=C['bg_header']))
    cell.border = accent_border
    for c in range(col_start + 1, col_end + 1):
        mc = ws.cell(row=row, column=c)
        mc.fill = px_fill(C['bg_page'])
        mc.border = accent_border









def _write_holdings(workbook, sheet, metrics: PortfolioMetrics):
    """Holdings: full enriched table with instrument type, data source, timestamp."""
    df = metrics.holdings_df
    if df.empty:
        _write_data_cell(sheet, 1, 1, "No holdings data", 0)
        return

    _apply_title(sheet, 1, 1, "Holdings Detail")

    columns = [
        ("Name", "name", False, None, False, False),
        ("Ticker", "ticker", False, None, False, False),
        ("ISIN", "isin", False, None, False, False),
        ("Asset Class", "asset_class", False, None, False, False),
        ("Security Type", "security_type", False, None, False, False),
        ("Currency", "currency", False, None, False, False),
        ("Quantity", "quantity", True, '#,##0.00', False, False),
        ("Avg Price", "avg_purchase_price", True, '#,##0.00', False, False),
        ("Current Price", "current_price", True, '#,##0.00', False, False),
        ("Cost Basis (EUR)", "cost_basis_eur", True, '#,##0.00', False, False),
        ("Value (EUR)", "current_value", True, '#,##0.00', False, False),
        ("% of Portfolio", "weight_pct", True, '0.00', False, False),
        ("% of Invested", "weight_of_invested_pctg", True, '0.00', False, False),
        ("% of Asset Class", "pct_of_class", True, '0.00', False, True),
        ("Gain (EUR)", "gain_eur", True, '#,##0.00', True, False),
        ("Gain %", "gain_pct", True, '0.00', True, False),
        ("Geography", "geography", False, None, False, False),
        ("Geo Source", "geo_source", False, None, False, False),
        ("Data Source", "data_source", False, None, False, False),
        ("Fetch Time", "fetch_timestamp", False, None, False, False),
    ]

    row = 3
    for c, (header, _, _, _, _, _) in enumerate(columns):
        _apply_header(sheet, row, c + 1, header)

    for idx, (_, data_row) in enumerate(df.iterrows()):
        row = idx + 4
        for c, (_, col_key, is_num, nf, use_gain_color, use_class_color) in enumerate(columns):
            val = data_row.get(col_key)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                _write_data_cell(sheet, row, c + 1, "", idx)
                continue

            # The Asset Class text column uses its own class-colored label.
            # The % of Asset Class numeric column keeps the numeric alignment
            # (center) but picks up the asset class color via font_color.
            ac = data_row.get("asset_class")
            if col_key == "asset_class":
                _write_data_cell(sheet, row, c + 1, val, idx,
                                 asset_class=ac, num_fmt=nf)
                continue

            if use_class_color and ac in ASSET_COLORS:
                class_color = ASSET_COLORS[ac]
                _write_data_cell(sheet, row, c + 1, val, idx,
                                 is_number=is_num, num_fmt=nf,
                                 font_color=class_color)
                continue

            # Gain columns keep the sign-based semaphore (green/red) via
            # _num_color. All other numeric columns render in neutral text.
            if use_gain_color:
                font_color = None
            elif is_num:
                font_color = C['text_pri']
            else:
                font_color = None
            _write_data_cell(sheet, row, c + 1, val, idx, is_number=is_num,
                             num_fmt=nf, font_color=font_color)

    _write_footer(sheet, row + 2, 1)



def _build_sensitivity_notes(sensitivity: list[dict], configured_w: float) -> list[str]:
    """Generate plain-language hints comparing the active drift-penalty
    regime to the immediate alternatives.

    The notes are designed to fit on one or two lines each so the user
    can skim them. They quantify the trade-off ("paying €X more in
    friction buys you Y pp less drift") rather than offering opinions.

    Returns an empty list when no actionable comparison exists, e.g.
    when there is only one regime in the sweep.
    """
    if not sensitivity or len(sensitivity) < 2:
        return []

    # Find the index of the active regime.
    active_idx = None
    for i, r in enumerate(sensitivity):
        if float(r["weight_min"]) <= configured_w <= float(r["weight_max"]):
            active_idx = i
            break
    if active_idx is None:
        return []

    from tarzan.export._format import eur_smart as _eur

    notes: list[str] = []
    active = sensitivity[active_idx]

    def _delta_summary(other: dict, direction: str) -> Optional[str]:
        a_friction = float(active["total_tax"]) + float(active["total_fee"])
        o_friction = float(other["total_tax"]) + float(other["total_fee"])
        d_friction = o_friction - a_friction
        d_drift_pp = float(other["max_drift_pp"]) - float(active["max_drift_pp"])
        d_actions = (other["n_buy"] + other["n_sell"]) - (active["n_buy"] + active["n_sell"])
        # Range label of the candidate regime.
        wmin = float(other["weight_min"])
        wmax = float(other["weight_max"])
        range_str = f"{wmin:g}" if wmin == wmax else f"{wmin:g}–{wmax:g}"
        if direction == "down":
            cost_word = "save" if d_friction < 0 else "spend"
            drift_word = "loosen" if d_drift_pp > 0 else "tighten"
            return (
                f"\u2193 Lowering to {range_str}: "
                f"{cost_word} {_eur(abs(d_friction))} in friction, "
                f"{drift_word} drift by {abs(d_drift_pp):.2f} pp, "
                f"{abs(d_actions):+d} actions."
            ).replace("+-", "−")
        else:
            cost_word = "spend" if d_friction > 0 else "save"
            drift_word = "tighten" if d_drift_pp < 0 else "loosen"
            return (
                f"\u2191 Raising to {range_str}: "
                f"{cost_word} {_eur(abs(d_friction))} more in friction, "
                f"{drift_word} drift by {abs(d_drift_pp):.2f} pp, "
                f"{abs(d_actions):+d} actions."
            ).replace("+-", "−")

    if active_idx > 0:
        prev = sensitivity[active_idx - 1]
        notes.append(_delta_summary(prev, "down"))
    if active_idx < len(sensitivity) - 1:
        nxt = sensitivity[active_idx + 1]
        notes.append(_delta_summary(nxt, "up"))
    return [n for n in notes if n]


def _write_allocations(workbook, sheet, metrics: PortfolioMetrics, config: InvestorConfig):
    """Optimizer: status overview, rebalancing actions, consolidated deviations table."""
    _apply_title(sheet, 1, 1, "Portfolio Optimizer")

    # Column widths. Both Allocation Deviations and Rebalancing Actions
    # tables now share the same 7-column layout (Ticker is in column B
    # for every row, kept empty for asset/geography buckets). The same
    # widths are reused for the Actions table where col G holds the
    # "Reason" string — that column is set to a wide enough size for
    # both purposes.
    sheet.column_dimensions['A'].width = 36   # Category / Holding name
    sheet.column_dimensions['B'].width = 13   # Ticker
    sheet.column_dimensions['C'].width = 13   # Current % (Allocation) | Direction (Actions)
    sheet.column_dimensions['D'].width = 16   # Target %               | Amount EUR
    sheet.column_dimensions['E'].width = 16   # Post-rebal %           | % of Portfolio
    sheet.column_dimensions['F'].width = 13   # Delta                  | Tax EUR
    sheet.column_dimensions['G'].width = 45   # Status                 | Reason
    # Column H is used by the drift-penalty sensitivity table for
    # the SELL-by-class summary; widen it so the EUR breakdown is
    # readable without truncation.
    sheet.column_dimensions['H'].width = 45

    tol = config.rebalancing_target_tolerance_pctg if config else 5.0

    # =====================================================================
    # OVERVIEW banner — traffic-light status based on largest deviation
    # =====================================================================
    max_abs_delta = 0.0
    if metrics.goal_deltas is not None and not metrics.goal_deltas.empty:
        # Exclude cash row: its delta_pct is relative to the cash target,
        # not an allocation percentage point. Cash is reported in EUR
        # further down in its own Cash Buffer section.
        non_cash = metrics.goal_deltas[metrics.goal_deltas["type"] != "cash"]
        if not non_cash.empty:
            max_abs_delta = float(non_cash["delta_pct"].abs().max())
    n_actions = len(metrics.rebalancing_suggestions) if metrics.rebalancing_suggestions else 0

    # When the engine returned 0 actions but at least one verification
    # row carries ``no_solution=True``, the LP was infeasible at the
    # configured tolerance ceiling — distinct from "already aligned".
    no_solution = bool(
        metrics.rebalancing_verifications
        and any(v.get("no_solution") for v in metrics.rebalancing_verifications)
    )
    # When the LP needed to relax the user-configured tolerance to find
    # a plan, surface this clearly so the user does not silently accept
    # a wider drift than they asked for.
    relaxed_meta = None
    if metrics.rebalancing_verifications:
        for v in metrics.rebalancing_verifications:
            if v.get("relaxed"):
                relaxed_meta = (
                    float(v.get("tolerance") or 0.0),
                    float(v.get("configured_max_tolerance") or tol),
                )
                break

    if no_solution:
        banner_color = C['red']
        banner_icon = "\u26a0"  # warning triangle
        max_tol = float(config.rebalancing_target_tolerance_pctg or 0.0)
        relax_cap = float(config.rebalancing_relax_cap_pctg or 0.0)
        if config.rebalancing_auto_relax and relax_cap > max_tol:
            banner_text = (
                f"No feasible rebalance even at the auto-relax cap "
                f"\u00b1{relax_cap:.1f}%. Targets are too aggressive — "
                f"review per-holding targets in the deviations table below."
            )
        else:
            banner_text = (
                f"No feasible rebalance — at least one allocation drift "
                f"exceeds \u00b1{max_tol:.1f}%. Consider raising max tolerance, "
                f"enabling rebalancing_auto_relax, or relaxing per-holding "
                f"targets."
            )
    elif n_actions == 0:
        # The LP solved cleanly and emitted no trades. Drift left is
        # either inside the band or pinned by locked positions —
        # either way, nothing actionable. Read as "Aligned" so the
        # banner does not contradict the empty actions table below.
        banner_color = C['green']
        banner_icon = "\u25cf"  # filled circle
        if max_abs_delta > 2 * tol:
            banner_text = (
                f"Aligned — largest reachable drift {max_abs_delta:.1f}% "
                f"is pinned by locked positions; nothing actionable."
            )
        elif max_abs_delta > tol:
            banner_text = (
                f"Aligned — drift of {max_abs_delta:.1f}% is within reach "
                f"of the tolerance band; no trades needed."
            )
        else:
            banner_text = f"Aligned — all allocations within \u00b1{tol:.1f}%"
    elif relaxed_meta is not None:
        used_tol, cfg_tol = relaxed_meta
        banner_color = C['amber']
        banner_icon = "\u26a0"
        banner_text = (
            f"Relaxed solution — solved at \u00b1{used_tol:.1f}% (configured "
            f"ceiling \u00b1{cfg_tol:.1f}%). Review per-holding targets if "
            f"this gap is uncomfortable."
        )
    elif max_abs_delta <= tol:
        banner_color = C['green']
        banner_icon = "\u25cf"  # filled circle
        banner_text = f"Aligned — all allocations within \u00b1{tol:.1f}%"
    elif max_abs_delta <= 2 * tol:
        banner_color = C['amber']
        banner_icon = "\u25cf"
        banner_text = (
            f"Minor drift — largest deviation {max_abs_delta:.1f}%"
            f" (tolerance \u00b1{tol:.1f}%). Rebalancing optional."
        )
    else:
        banner_color = C['red']
        banner_icon = "\u25cf"
        banner_text = (
            f"Action needed — largest deviation {max_abs_delta:.1f}%"
            f" (>{2 * tol:.1f}%). Rebalancing recommended."
        )
    if n_actions > 0:
        banner_text += f"  \u00b7  {n_actions} action{'s' if n_actions != 1 else ''} suggested below"

    row = 3
    bcell = sheet.cell(row=row, column=1, value=f"{banner_icon}  {banner_text}")
    bcell.font = px_font(size=11, bold=True, color=banner_color)
    bcell.fill = px_fill(C['bg_page'])
    bcell.alignment = px_align(h='left', v='center')
    bcell.border = px_no_border()
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    sheet.row_dimensions[row].height = 22
    row += 2

    # =====================================================================
    # REBALANCING ACTIONS
    # =====================================================================
    header_parts = ["REBALANCING ACTIONS"]
    if config.rebalancing_lump_sum_amount_eur > 0:
        header_parts.append(
            f"lump sum {_format_number(config.rebalancing_lump_sum_amount_eur)} EUR"
        )
    if config.rebalancing_no_sell:
        header_parts.append("no-sell mode")
    _write_area_header(sheet, row, 1, 7, " \u00b7 ".join(header_parts))
    row += 1

    if metrics.rebalancing_suggestions:
        # Compute totals + capital-gains tax / fee impact in the same way
        # the LP solver applies them, so the summary line matches the
        # solver's view of friction.
        cg_std = float(config.rebalancing_capital_gains_tax_standard_pctg or 0.0) / 100.0
        cg_gov = float(config.rebalancing_capital_gains_tax_government_pctg or 0.0) / 100.0
        fee_buy = float(config.rebalancing_transaction_fee_buy_eur or 0.0)
        fee_sell = float(config.rebalancing_transaction_fee_sell_eur or 0.0)

        total_buy = sum(s["amount_eur"] for s in metrics.rebalancing_suggestions if s["direction"] == "buy")
        total_sell = sum(s["amount_eur"] for s in metrics.rebalancing_suggestions if s["direction"] == "sell")

        # Per-row tax = gross × max(0, gain_pct/100) × cg_rate. Used both
        # for the per-row "Tax (EUR)" column and for the summary total.
        df = metrics.holdings_df
        per_row_tax: dict[str, float] = {}
        total_tax = 0.0
        for s in metrics.rebalancing_suggestions:
            if s["direction"] != "sell":
                per_row_tax[s["ticker"]] = 0.0
                continue
            row_df = df[df["ticker"] == s["ticker"]]
            if row_df.empty:
                per_row_tax[s["ticker"]] = 0.0
                continue
            hr = row_df.iloc[0]
            gain = float(hr.get("gain_pct") or 0.0)
            if gain <= 0:
                per_row_tax[s["ticker"]] = 0.0
                continue
            instr = (hr.get("instrument_type") or "").lower()
            rate = cg_gov if "government bond" in instr else cg_std
            t = s["amount_eur"] * (gain / 100.0) * rate
            per_row_tax[s["ticker"]] = t
            total_tax += t

        n_buy = sum(1 for s in metrics.rebalancing_suggestions if s["direction"] == "buy")
        n_sell = sum(1 for s in metrics.rebalancing_suggestions if s["direction"] == "sell")
        total_fee = n_buy * fee_buy + n_sell * fee_sell

        summary_parts = [
            f"Total BUY: {_format_number(total_buy)} EUR",
            f"Total SELL: {_format_number(total_sell)} EUR",
            f"Net: {_format_number(total_buy - total_sell)} EUR",
        ]
        if total_tax > 0:
            summary_parts.append(f"Tax: {_format_number(total_tax)} EUR")
        if total_fee > 0:
            summary_parts.append(f"Fees: {_format_number(total_fee)} EUR")
        summary = "  \u00b7  ".join(summary_parts)
        scell = sheet.cell(row=row, column=1, value=summary)
        scell.font = px_font(size=10, bold=True, color=C['text_sec'])
        scell.fill = px_fill(C['bg_page'])
        scell.alignment = px_align(h='left')
        scell.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 2

        for c, h in enumerate(
            ["Holding", "Ticker", "Direction", "Amount (EUR)", "% of Portfolio", "Tax (EUR)", "Reason"], 1
        ):
            _apply_header(sheet, row, c, h)
        row += 1
        total_value = metrics.total_value or 1.0
        for ti, s in enumerate(metrics.rebalancing_suggestions):
            direction = s["direction"].upper()
            pct_of_port = (s["amount_eur"] / total_value) * 100 if total_value > 0 else 0
            dir_color = C['green'] if direction == "BUY" else C['red']
            tax_value = per_row_tax.get(s["ticker"], 0.0)
            _write_data_cell(sheet, row, 1, s.get("name", ""), ti)
            _write_data_cell(sheet, row, 2, s.get("ticker", ""), ti, font_color=C['text_sec'])
            _write_data_cell(sheet, row, 3, direction, ti, bold=True, font_color=dir_color)
            _write_data_cell(sheet, row, 4, s["amount_eur"], ti, is_number=True,
                             num_fmt='#,##0.00', font_color=C['text_pri'])
            _write_data_cell(sheet, row, 5, pct_of_port, ti, is_number=True,
                             num_fmt='0.00', font_color=C['text_pri'])
            # Always write a numeric value so the column reads
            # consistently — show "0.00" for buys and for sells of
            # positions in loss (where realising the loss generates
            # no capital gains tax). The legend in the dashboard makes
            # it clear that 0 means "no taxable gain".
            _write_data_cell(sheet, row, 6, tax_value, ti, is_number=True,
                             num_fmt='#,##0.00', font_color=C['text_pri'])
            _write_data_cell(sheet, row, 7, s.get("reason", ""), ti)
            row += 1
    else:
        if no_solution:
            max_tol = float(config.rebalancing_target_tolerance_pctg or 0.0)
            no_msg = (
                f"No feasible solution within \u00b1{max_tol:.1f}% tolerance. "
                f"At least one allocation drift exceeds the ceiling — "
                f"raise rebalancing_target_tolerance_pctg or relax per-holding "
                f"targets to obtain a plan."
            )
        else:
            no_msg = (
                "Portfolio already aligned within tolerance — no actions needed."
            )
        nocell = sheet.cell(row=row, column=1, value=no_msg)
        nocell.font = px_font(size=10, italic=True, color=C['text_sec'])
        nocell.fill = px_fill(C['bg_page'])
        nocell.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

    # =====================================================================
    # ALLOCATION DEVIATIONS — grouped by type, with post-rebalancing view
    # =====================================================================
    row += 2
    _write_area_header(sheet, row, 1, 7, "ALLOCATION DEVIATIONS")
    row += 1
    subcell_text = (
        f"Status color based on delta after rebalancing vs target. "
        f"Threshold \u00b1{tol:.1f}% (green), \u00b1{2 * tol:.1f}% (amber), beyond (red)."
    )
    if no_solution:
        subcell_text = (
            f"No feasible rebalance — Post-rebal column omitted. "
            f"Status color based on the current drift vs target. "
            f"Threshold \u00b1{tol:.1f}% (green), \u00b1{2 * tol:.1f}% (amber), beyond (red)."
        )
    subcell = sheet.cell(row=row, column=1, value=subcell_text)
    subcell.font = px_font(size=9, italic=True, color=C['text_sec'])
    subcell.fill = px_fill(C['bg_page'])
    subcell.border = px_no_border()
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    # Build current (goal_deltas) lookup by (type, category). Skip cash
    # rows: cash uses EUR fields and a dedicated section.
    current_lookup: dict[tuple[str, str], tuple[float, float]] = {}
    if metrics.goal_deltas is not None and not metrics.goal_deltas.empty:
        for _, gd in metrics.goal_deltas.iterrows():
            if gd["type"] == "cash":
                continue
            tp = "asset" if gd["type"] == "asset_class" else "geography"
            current_lookup[(tp, gd["category"])] = (
                float(gd["actual_pct"]), float(gd["target_pct"]),
            )

    # Per-holding pre-rebalancing lookup, keyed by (kind, ticker). The
    # ``actual_pct`` column in the verifications is post-rebalancing, so
    # we recompute the pre-rebalancing within-class percentage from
    # ``holdings_df`` here. ``ticker`` is used as the key (instead of
    # name) because two holdings can share a name — for example a user
    # holding two BTPs both labelled "BUONI POLIENNALI DEL TES" — but
    # ticker is unique. Without this, the table showed identical values
    # in the Current and Post-rebal columns for every per-holding row.
    per_holding_pre_lookup: dict[tuple[str, str], float] = {}
    df = metrics.holdings_df
    if df is not None and not df.empty and "pct_of_class" in df.columns:
        for _, hr in df.iterrows():
            ac = hr.get("asset_class")
            if ac == "Equities":
                per_holding_pre_lookup[("per_holding_equity", hr["ticker"])] = float(hr["pct_of_class"])
            elif ac == "Fixed Income":
                per_holding_pre_lookup[("per_holding_fi", hr["ticker"])] = float(hr["pct_of_class"])

    # Build post-rebalancing lookup by (kind, key). Skip cash — its items
    # use EUR fields and are rendered in the Cash Buffer section. Asset
    # and Geography kinds key off the bucket name. Per-holding kinds key
    # off ``ticker`` to avoid collisions when two holdings share a name.
    post_lookup: dict[tuple[str, str], tuple[float, float]] = {}
    # Per-holding rows also need a display label (the holding name) and
    # the original verification entry order is preserved for stable sort.
    per_holding_meta: dict[tuple[str, str], dict] = {}
    if metrics.rebalancing_verifications:
        for v in metrics.rebalancing_verifications:
            if v.get("kind") == "cash":
                continue
            for it in v.get("items", []) or []:
                key_id = it.get("ticker") if v["kind"].startswith("per_holding_") else it["category"]
                post_lookup[(v["kind"], key_id)] = (
                    float(it["actual_pct"]), float(it["target_pct"]),
                )
                if v["kind"].startswith("per_holding_"):
                    per_holding_meta[(v["kind"], key_id)] = {
                        "name": it["category"],
                    }

    # Define the four groups in the desired order. Cash buffer is merged
    # into the first (asset) group as an EUR-denominated row.
    groups = [
        ("asset", "Invested Allocation", current_lookup),
        ("geography", "Geography (equity only)", current_lookup),
        ("per_holding_equity", "Per-Holding Equity Targets", None),
        ("per_holding_fi", "Per-Holding Fixed Income Targets", None),
    ]

    # Pull cash numbers once so we can inject a row in the asset group.
    cash_actual = metrics.cash_value
    cash_target_eur = metrics.cash_target_eur
    cash_post = cash_actual
    if metrics.rebalancing_verifications:
        for v in metrics.rebalancing_verifications:
            if v.get("kind") == "cash" and v.get("items"):
                cash_post = float(v["items"][0].get("actual_eur", cash_actual))
                break
    cash_delta_eur = cash_post - cash_target_eur

    for kind, title, current_source in groups:
        # Collect categories: for asset/geography we use bucket names,
        # for per-holding kinds we use tickers (and resolve the display
        # name later from per_holding_meta).
        categories: list[str] = []
        if kind in ("asset", "geography"):
            categories = [cat for (tp, cat) in current_lookup if tp == kind]
        else:
            for (k, cat) in post_lookup:
                if k == kind:
                    categories.append(cat)

        # Skip group entirely if no categories AND (for asset) no cash row to add.
        if not categories and not (kind == "asset" and cash_target_eur > 0):
            continue

        # Group header
        hdr = sheet.cell(row=row, column=1, value=title)
        hdr.font = px_font(size=10, bold=True, color=C['text_pri'])
        hdr.fill = px_fill(C['bg_page'])
        hdr.alignment = px_align(h='left')
        hdr.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # Column headers. All four sub-tables share the same 7-column
        # layout so the numeric columns line up vertically across
        # groups (asset class, geography, per-holding equity, per-
        # holding FI). The Ticker column is empty for asset/geo rows
        # and populated for per-holding rows. Header label switches
        # between "Category" and "Holding" depending on the section.
        is_per_holding = kind.startswith("per_holding_")
        row_label_header = "Holding" if is_per_holding else "Category"
        col_headers = [
            row_label_header, "Ticker",
            "Current (% / EUR)", "Target (% / EUR)",
            "Post-rebal (% / EUR)", "Delta (pp / EUR)", "Status",
        ]
        for c, h in enumerate(col_headers, 1):
            _apply_header(sheet, row, c, h)
        row += 1

        # Sort categories by |post-rebal delta|, then by |current delta|, descending
        def _sort_key(cat: str) -> float:
            post = post_lookup.get((kind, cat))
            if post:
                return abs(post[0] - post[1])
            curr = current_source.get((kind, cat)) if current_source else None
            return abs(curr[0] - curr[1]) if curr else 0.0

        categories.sort(key=_sort_key, reverse=True)

        for ti, cat in enumerate(categories):
            # Current (pre-rebalancing). For asset/geography we use the
            # goal_deltas snapshot; for per-holding rows we use the
            # holdings_df pct-of-class column captured in
            # per_holding_pre_lookup. ``cat`` is a bucket name for asset
            # and geography, and a ticker for per-holding kinds.
            curr_tuple = current_source.get((kind, cat)) if current_source else None
            current_pct = curr_tuple[0] if curr_tuple else None
            if current_pct is None and kind.startswith("per_holding_"):
                current_pct = per_holding_pre_lookup.get((kind, cat))
            # Target + post from verifications (authoritative for per-holding)
            post_tuple = post_lookup.get((kind, cat))
            if post_tuple:
                post_pct, target_pct = post_tuple
            else:
                post_pct = None
                target_pct = curr_tuple[1] if curr_tuple else None

            # Last-resort fallback: still no current after both sources
            # have been tried. Use the post value so the column is not
            # blank, which keeps the row legible even when something is
            # missing from holdings_df (very rare).
            if current_pct is None and post_pct is not None:
                current_pct = post_pct

            # Resolve display label: bucket name as-is for asset/geo,
            # the holding name (resolved from per_holding_meta) for
            # per-holding rows.
            if kind.startswith("per_holding_"):
                display_label = per_holding_meta.get((kind, cat), {}).get("name", cat)
            else:
                display_label = cat

            # Delta logic
            #   * normal case: post_pct/target_pct both available → use
            #     post-rebal delta to drive color and status.
            #   * no-solution case: actions=[] and the LP was infeasible
            #     → post_pct equals current_pct (no rebalance happened).
            #     We hide the Post-rebal column and report Delta from the
            #     current values so the user sees the live drift.
            if no_solution:
                post_pct_display = None
                if current_pct is not None and target_pct is not None:
                    delta_after = current_pct - target_pct
                else:
                    delta_after = 0.0
            else:
                post_pct_display = post_pct
                if post_pct is not None and target_pct is not None:
                    delta_after = post_pct - target_pct
                elif current_pct is not None and target_pct is not None:
                    delta_after = current_pct - target_pct
                else:
                    delta_after = 0.0

            color = _deviation_color(delta_after, tol)
            abs_d = abs(delta_after)
            if abs_d <= tol:
                status = "\u25cf Aligned"
            elif abs_d <= 2 * tol:
                status = "\u25cf Drift"
            else:
                status = "\u25cf Action"

            # Match Dashboard look: color the Category label with the asset
            # class or geography palette when applicable. For per-holding
            # tables the row label is the holding name (no palette tag)
            # and the Ticker cell carries the ticker. For asset/geo
            # rows the Ticker cell is blank but kept so that numeric
            # columns line up across all four sub-tables.
            if kind == "asset":
                _write_data_cell(sheet, row, 1, display_label, ti, asset_class=display_label)
            elif kind == "geography":
                _write_data_cell(sheet, row, 1, display_label, ti, geography=display_label)
            else:
                _write_data_cell(sheet, row, 1, display_label, ti)

            # Ticker cell. Always written so column B has consistent
            # styling across rows; empty string for asset/geo, the
            # actual ticker for per-holding rows.
            ticker_value = cat if is_per_holding else ""
            _write_data_cell(sheet, row, 2, ticker_value, ti, font_color=C['text_sec'])
            col_current, col_target, col_post, col_delta, col_status = 3, 4, 5, 6, 7

            _write_data_cell(sheet, row, col_current,
                             current_pct if current_pct is not None else "",
                             ti, is_number=current_pct is not None,
                             num_fmt='0.00' if current_pct is not None else None,
                             font_color=C['text_pri'])
            _write_data_cell(sheet, row, col_target,
                             target_pct if target_pct is not None else "",
                             ti, is_number=target_pct is not None,
                             num_fmt='0.00' if target_pct is not None else None,
                             font_color=C['text_pri'])
            _write_data_cell(sheet, row, col_post,
                             post_pct_display if post_pct_display is not None else "—",
                             ti, is_number=post_pct_display is not None,
                             num_fmt='0.00' if post_pct_display is not None else None,
                             font_color=C['text_pri'])
            _write_data_cell(sheet, row, col_delta, delta_after, ti, is_number=True,
                             num_fmt='+0.00;-0.00;0.00', font_color=color)
            _write_data_cell(sheet, row, col_status, status, ti, bold=True, font_color=color)
            row += 1

        # Append the Cash Buffer row at the bottom of the asset group.
        if kind == "asset" and cash_target_eur > 0:
            rel_dev = (cash_delta_eur / cash_target_eur) * 100.0
            cash_color = _deviation_color(rel_dev, tol)
            abs_rel = abs(rel_dev)
            if abs_rel <= tol:
                cash_status = "\u25cf Aligned"
            elif abs_rel <= 2 * tol:
                cash_status = "\u25cf Drift"
            else:
                cash_status = "\u25cf Action"
            ti2 = len(categories)
            # Cash row uses the same 7-column layout: empty Ticker
            # (col 2) keeps the numeric columns lined up with the
            # rest of the Allocation Deviations section.
            _write_data_cell(sheet, row, 1, "Cash & Cash Equivalents", ti2,
                             asset_class="Cash & Cash Equivalents")
            _write_data_cell(sheet, row, 2, "", ti2)
            _write_data_cell(sheet, row, 3, cash_actual, ti2, is_number=True,
                             num_fmt='"€"#,##0.00', font_color=C['text_pri'])
            _write_data_cell(sheet, row, 4, cash_target_eur, ti2, is_number=True,
                             num_fmt='"€"#,##0.00', font_color=C['text_pri'])
            # Hide the Post-rebal cash value when the LP was infeasible
            # (it would equal the current value and falsely suggest a
            # rebalance happened).
            if no_solution:
                _write_data_cell(sheet, row, 5, "—", ti2)
            else:
                _write_data_cell(sheet, row, 5, cash_post, ti2, is_number=True,
                                 num_fmt='"€"#,##0.00', font_color=C['text_pri'])
            _write_data_cell(sheet, row, 6, cash_delta_eur, ti2, is_number=True,
                             num_fmt='"€"+#,##0.00;"€"-#,##0.00;"€"0.00',
                             font_color=cash_color)
            _write_data_cell(sheet, row, 7, cash_status, ti2, bold=True, font_color=cash_color)
            row += 1

        row += 1  # spacer between groups

    if not current_lookup and not post_lookup:
        nocell = sheet.cell(row=row, column=1, value="No targets configured.")
        nocell.font = px_font(size=10, italic=True, color=C['text_sec'])
        nocell.fill = px_fill(C['bg_page'])
        nocell.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

    # =====================================================================
    # DRIFT-PENALTY SENSITIVITY — show the user how the optimization
    # changes as ``rebalancing_drift_penalty_weight`` varies, so they
    # can pick the value that matches their preference (more trades
    # vs. less leftover drift).
    # =====================================================================
    sensitivity = getattr(metrics, "rebalancing_sensitivity", None)
    # Hide the sensitivity card when every regime produces zero
    # trades — typical when the portfolio is already inside the
    # tolerance band at every weight. Otherwise the section ships an
    # empty diagnostic that wastes vertical space.
    has_trades = bool(sensitivity) and any(
        (r.get("n_buy", 0) + r.get("n_sell", 0)) > 0
        for r in sensitivity
    )
    if sensitivity and has_trades:
        from tarzan.export._format import eur_smart as _eur_compact
        row += 2
        _write_area_header(sheet, row, 1, 7, "DRIFT-PENALTY SENSITIVITY")
        row += 1
        configured_w = float(getattr(config, "rebalancing_drift_penalty_weight", 0.0) or 0.0)
        subcell_text = (
            f"Each row is a regime — a range of weights producing the "
            f"same plan. Active weight: {configured_w:g} (highlighted with \u25b6). "
            f"Higher weights close more drift at the cost of more trades, "
            f"taxes, and fees."
        )
        subcell = sheet.cell(row=row, column=1, value=subcell_text)
        subcell.font = px_font(size=9, italic=True, color=C['text_sec'])
        subcell.fill = px_fill(C['bg_page'])
        subcell.border = px_no_border()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 2

        sens_headers = [
            "Weight range",
            "Actions (B/S)",
            "BUY total",
            "SELL total",
            "Tax + Fee",
            "Max drift",
            "Net change by class",
        ]
        for c, h in enumerate(sens_headers, 1):
            _apply_header(sheet, row, c, h)
        row += 1

        def _net_by_class_str(d: dict, min_eur: float = 0.5) -> str:
            """Render a class-net dict as ``Equ +€1.9k  ·  FI −€20.1k``.

            Sorted by descending absolute volume; entries below
            ``min_eur`` are dropped (numerical noise from the LP
            solver). Sign comes from BUY − SELL: positive values mean
            the class grows, negative shrinks.
            """
            items = sorted((d or {}).items(), key=lambda x: -abs(x[1]))
            return "  \u00b7  ".join(
                f"{ac} {_eur_compact(v, signed=True)}"
                for ac, v in items
                if abs(v) > min_eur
            ) or "—"

        for ti, regime in enumerate(sensitivity):
            wmin = float(regime["weight_min"])
            wmax = float(regime["weight_max"])
            range_str = f"{wmin:g}" if wmin == wmax else f"{wmin:g} – {wmax:g}"
            actions_str = f"{regime['n_buy']}B / {regime['n_sell']}S"
            buy = float(regime["total_buy"])
            sell = float(regime["total_sell"])
            friction = float(regime["total_tax"]) + float(regime["total_fee"])
            drift = float(regime["max_drift_pp"])
            net_str = _net_by_class_str(regime.get("net_by_class") or {})

            is_active = (wmin <= configured_w <= wmax)
            label_color = C['text_pri'] if is_active else C['text_sec']
            range_value = f"\u25b6 {range_str}" if is_active else range_str
            _write_data_cell(sheet, row, 1, range_value, ti, bold=is_active,
                             font_color=label_color)
            _write_data_cell(sheet, row, 2, actions_str, ti, font_color=label_color)
            _write_data_cell(sheet, row, 3, _eur_compact(buy), ti,
                             font_color=label_color)
            _write_data_cell(sheet, row, 4, _eur_compact(sell), ti,
                             font_color=label_color)
            _write_data_cell(sheet, row, 5, _eur_compact(friction), ti,
                             font_color=label_color)
            _write_data_cell(sheet, row, 6, drift, ti, is_number=True,
                             num_fmt='0.00"%"', font_color=label_color)
            _write_data_cell(sheet, row, 7, net_str, ti, font_color=label_color)
            row += 1

        # Dynamic insight: compare the active regime to one notch up
        # and one notch down so the user sees what changing the weight
        # would buy them.
        notes = _build_sensitivity_notes(sensitivity, configured_w)
        if notes:
            row += 1
            for note in notes:
                ncell = sheet.cell(row=row, column=1, value=note)
                ncell.font = px_font(size=10, italic=True, color=C['text_sec'])
                ncell.fill = px_fill(C['bg_page'])
                ncell.alignment = px_align(h='left', wrap=True)
                ncell.border = px_no_border()
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                row += 1

    # =====================================================================
    # SOLVER INFO
    # =====================================================================
    row += 2
    _write_area_header(sheet, row, 1, 7, "SOLVER PARAMETERS")
    row += 1

    tol_used = None
    if metrics.rebalancing_verifications:
        tol_used = metrics.rebalancing_verifications[0].get("tolerance")

    info_rows = [
        ("Target tolerance",
         f"\u00b1{config.rebalancing_target_tolerance_pctg:.1f}%",
         "Tolerance band around every allocation target. The LP uses it "
         "as the hard ceiling and the dashboard uses it as the "
         "traffic-light threshold (green ≤ value, amber up to 2×, red beyond)."),
        ("Solver tolerance",
         (f"\u00b1{tol_used:.2f}%" if tol_used is not None else "n/a"),
         "Actual tolerance the optimizer converged at (progressive up to target)"),
        ("Lump sum",
         (f"{_format_number(config.rebalancing_lump_sum_amount_eur)} EUR"
          if config.rebalancing_lump_sum_amount_eur > 0 else "—"),
         "Additional cash to deploy in the rebalance"),
        ("Cash & Cash Equivalents target",
         (f"{_format_number(config.target_cash_buffer_eur)} EUR"
          if config.target_cash_buffer_eur > 0 else "—"),
         "Absolute cash target used by the solver (from target_cash_buffer_eur)"),
        ("No-sell mode",
         ("enabled" if config.rebalancing_no_sell else "disabled"),
         "If enabled, the solver can only buy, never sell"),
        ("Drift-penalty weight",
         f"{getattr(config, 'rebalancing_drift_penalty_weight', 0.0):g}",
         "Cost in the LP objective for residual EUR drift outside the "
         "tolerance band. 0 = pure minimum trading. 1 = balanced (default). "
         "Higher values close out-of-band drift more aggressively."),
    ]

    for c, h in enumerate(["Parameter", "Value", "", "", "", "", "Description"], 1):
        if h:
            _apply_header(sheet, row, c, h)
    row += 1
    for ti, (param, value, desc) in enumerate(info_rows):
        _write_data_cell(sheet, row, 1, param, ti, bold=True)
        _write_data_cell(sheet, row, 2, value, ti)
        _write_data_cell(sheet, row, 7, desc, ti)
        row += 1

    row += 1
    _write_footer(sheet, row, 1)





def _write_performance(workbook, sheet, metrics: PortfolioMetrics):
    """Performance: unified table with TOTAL PORTFOLIO + holdings + benchmarks.

    All risk metrics (CAGR, Vol, Sharpe, Sortino, Max DD, Alpha, Beta) are computed
    on the full available history per instrument, capped at 5 years. The Period Used
    column shows the actual window used.

    Alpha and Beta are computed vs the benchmark marked is_benchmark_alfa_and_beta=true
    in indexes.csv (column headers are dynamic).
    """
    _apply_title(sheet, 1, 1, "Performance Analysis")

    # Dynamic header for Alpha/Beta based on configured benchmark
    bench_beta_name = cfg.benchmark_beta_name()
    alpha_label = f"\u03b1 (vs {bench_beta_name})"
    beta_label = f"\u03b2 (vs {bench_beta_name})"

    # Info note
    row = 2
    note = (f"Period returns (1D–5Y) and risk metrics calculated on available history "
            f"per instrument (max 5 years). \u03b1/\u03b2 computed vs {bench_beta_name}.")
    cell = sheet.cell(row=row, column=1, value=note)
    cell.font = px_font(size=9, italic=True, color=C['text_sec'])
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=21)
    row = 4

    # Warning banner: holdings excluded from the TOTAL PORTFOLIO time series
    # because their price history span is below the 1Y threshold. Each row
    # is still shown individually below in its own row with its real Period
    # Used.
    excluded = getattr(metrics, "excluded_short_tenure", []) or []
    if excluded:
        names = ", ".join(item.get("name", item.get("ticker", "?")) for item in excluded)
        weight = sum(float(item.get("weight_pct", 0.0)) for item in excluded)
        warn_text = (
            f"\u26a0  TOTAL PORTFOLIO excludes {len(excluded)} holding"
            f"{'s' if len(excluded) > 1 else ''} with <1Y of price history "
            f"({weight:.1f}% of AuM): {names}. Each excluded holding is still "
            f"shown in its own row below."
        )
        warn_cell = sheet.cell(row=row, column=1, value=warn_text)
        warn_cell.font = px_font(size=9, italic=True, color=C['amber'])
        warn_cell.alignment = px_align(h='left', wrap=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=21)
        sheet.row_dimensions[row].height = 28
        row += 2

    # Returns coverage note: when an order list drove XIRR/TWROR, disclose
    # how much of the value rested on real market data and which
    # instruments were priced by the synthetic/carry-flat fallback.
    if metrics.returns_coverage_pct is not None:
        prov = metrics.returns_provenance or {}
        fallback = sorted(
            set(prov.get("synthetic", []))
            | set(prov.get("carry_flat", []))
            | set(prov.get("excluded", []))
        )
        cov_text = (
            f"XIRR {metrics.xirr_pct:+.2f}% (money-weighted)  ·  "
            f"TWROR {metrics.twror_pct:+.2f}% cumulative "
            f"({metrics.twror_annualized_pct:+.2f}% annualized).  "
            f"Computed on {metrics.returns_coverage_pct:.1f}% real market data."
        )
        if fallback:
            cov_text += (
                f"  {len(fallback)} instrument"
                f"{'s' if len(fallback) > 1 else ''} priced by fallback "
                f"(interpolated/carry-flat): {', '.join(fallback)}."
            )
        cov_cell = sheet.cell(row=row, column=1, value=cov_text)
        cov_cell.font = px_font(size=9, italic=True, color=C['text_sec'])
        cov_cell.alignment = px_align(h='left', wrap=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=21)
        sheet.row_dimensions[row].height = 28
        row += 2

    # Net-of-tax estimate disclosure: states the figure and the method, so
    # the estimate is transparent and never confused with the gross return.
    if (metrics.estimated_cgt_eur is not None and metrics.estimated_cgt_eur > 0):
        parts = [
            f"Net-of-tax estimate: \u2212\u20ac{metrics.estimated_cgt_eur:,.0f} "
            f"estimated CGT on realized gains"
        ]
        if metrics.pnl_eur_net_tax is not None:
            parts.append(f"PnL {metrics.pnl_eur_net_tax:+,.0f} EUR")
        if metrics.xirr_net_tax_pct is not None:
            parts.append(f"XIRR {metrics.xirr_net_tax_pct:+.2f}% ann")
        tax_text = (
            "  \u00b7  ".join(parts)
            + ".  Estimate only: average-cost basis, 26% / 12.5% on government "
            "bonds, realized losses offset later gains where Italian rules allow "
            "(ETF/fund gains are redditi di capitale, not offsettable). Excludes "
            "coupon/dividend withholding and the original cost basis of "
            "transferred-in positions. TWROR is shown gross of tax."
        )
        tax_cell = sheet.cell(row=row, column=1, value=tax_text)
        tax_cell.font = px_font(size=9, italic=True, color=C['text_sec'])
        tax_cell.alignment = px_align(h='left', wrap=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=21)
        sheet.row_dimensions[row].height = 40
        row += 2

    # Columns
    all_cols = ["1d", "1w", "1m", "3m", "6m", "ytd", "1y", "3y", "5y",
                "cagr", "volatility", "sharpe", "sortino", "max_drawdown",
                "var_95", "cvar_95", "alpha", "beta", "period_used"]
    all_labels = ["1D", "1W", "1M", "3M", "6M", "YTD", "1Y", "3Y", "5Y",
                  "CAGR", "Volatility", "Sharpe", "Sortino", "Max DD",
                  "VaR 95%", "CVaR 95%", alpha_label, beta_label, "Period Used"]

    # Set column widths
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 14
    for i in range(3, 21):
        sheet.column_dimensions[get_column_letter(i)].width = 11
    sheet.column_dimensions[get_column_letter(21)].width = 12

    # Header row
    _apply_header(sheet, row, 1, "Name")
    _apply_header(sheet, row, 2, "Type")
    for c, label in enumerate(all_labels, 3):
        _apply_header(sheet, row, c, label)
    row += 1

    # --- TOTAL PORTFOLIO row ---
    port_full = metrics.performance_full or {}
    port_vals = ["Portfolio"]
    for key in all_cols:
        val = port_full.get(key)
        port_vals.append(_format_number(val, True) if key != "period_used" else (val or "—"))
    _write_portfolio_row(sheet, row, 1, "** TOTAL PORTFOLIO **", port_vals, 0)
    row += 1

    # --- Holdings rows + Benchmarks rows ---
    if not metrics.holding_performance.empty:
        hp = metrics.holding_performance
        # Sort: In portfolio first, then Benchmark index
        hp_sorted = hp.sort_values(by="type", ascending=True, kind="stable")
        for ti, (_, hr) in enumerate(hp_sorted.iterrows()):
            _write_data_cell(sheet, row, 1, hr.get("name", hr.get("ticker", "")), ti)
            _write_data_cell(sheet, row, 2, hr.get("type", ""), ti)
            for c, key in enumerate(all_cols, 3):
                val = hr.get(key)
                if key == "period_used":
                    _write_data_cell(sheet, row, c, val or "—", ti)
                else:
                    display = _format_number(val, True)
                    _write_data_cell(sheet, row, c, display, ti,
                                     is_number=isinstance(display, (int, float)))
            row += 1

    row += 2

    # --- LEGEND: Rating Thresholds ---
    _apply_subtitle(sheet, row, 1, "Legend — Rating Thresholds")
    row += 1

    legend_headers = [
        "Metric", "\u25cf Strong", "\u25cf Fair", "\u25cf Weak", "Source", "Description",
    ]
    # Column widths: keep Metric/Strong/Fair/Weak/Source compact; give
    # Description a very wide column so it fits on a single line without
    # truncation.
    sheet.column_dimensions['A'].width = 20  # Metric
    for i in range(2, 5):
        sheet.column_dimensions[get_column_letter(i)].width = 18  # Strong / Fair / Weak
    sheet.column_dimensions['E'].width = 34  # Source
    sheet.column_dimensions['F'].width = 130  # Description (single line)
    for c, h in enumerate(legend_headers, 1):
        _apply_header(sheet, row, c, h)
    row += 1

    ratings = cfg.metric_ratings() or {}
    legend_rows = [
        ("CAGR", "cagr",
         "Compound Annual Growth Rate. The single yearly return that, if repeated every year, "
         "would grow your portfolio from start to end value. Accounts for compounding. "
         "~7% is the long-term global equity average.",
         "Equity risk premium (Dimson et al.)"),
        (alpha_label, "alpha",
         "Extra annual return vs the benchmark, after adjusting for how risky the portfolio is (CAPM). "
         "Positive = you beat the market beyond what your risk alone justified. "
         "Negative = you underperformed after fees and noise.",
         "Jensen's Alpha (CAPM)"),
        (beta_label, "beta",
         "How much your portfolio moves when the benchmark moves 1%. β=1 in line, β=0.5 half as "
         "reactive, β=1.5 amplifies by 50%, β≈0 uncorrelated. Tells you how much systematic "
         "market risk you're running.",
         "CAPM, 1.0 = market"),
        ("Max DD", "max_drawdown",
         "Maximum Drawdown. Worst peak-to-trough loss over the period — the most painful scenario "
         "an investor lived through. -20% is typical for diversified equity; deeper drops signal "
         "concentration or high volatility.",
         "Retail drawdown tolerance"),
        ("Volatility", "volatility",
         "How bumpy the ride is. Annualized standard deviation of daily returns. A 15% vol means "
         "~±15% year-to-year noise around the average return. Equity indexes ~15–20%, bonds ~3–7%.",
         "Hist. equity vol ~15%"),
        ("Sharpe", "sharpe",
         "Return per unit of risk taken: (CAGR − risk-free rate) / Volatility. "
         "Above 1 is good, above 2 is excellent, negative means you were paid less than a safe "
         "bond for the risk you ran.",
         "Sharpe (1994)"),
        ("Sortino", "sortino",
         "Like Sharpe but only penalizes downside volatility (ignores upside swings). More honest "
         "when returns are asymmetric. Usually higher than Sharpe — the gap shows how much of "
         "your volatility is actually good volatility.",
         "Sortino & Price (1994)"),
        ("VaR 95%", "var_pct",
         "Value at Risk. The daily loss exceeded only 5% of the time (historical simulation). "
         "A VaR of -1.2% means on an average month you should expect about one day worse than "
         "-1.2%. Non-parametric: no normal-distribution assumption.",
         "Basel III (retail adj.)"),
        ("CVaR 95%", "cvar_pct",
         "Conditional VaR, a.k.a. Expected Shortfall. The average loss on the worst 5% of days. "
         "Always more negative than VaR. Captures tail risk VaR misses — how bad it really gets "
         "when it goes bad.",
         "Artzner et al. (1999)"),
    ]
    for ti, (metric_label, key, description, source) in enumerate(legend_rows):
        spec = ratings.get(key, {})
        thresholds = spec.get("thresholds", [None, None])
        labels = spec.get("labels", ["—", "—", "—"])
        invert = spec.get("invert", False)
        unit = spec.get("unit", "")
        good_t, warn_t = thresholds[0], thresholds[1]

        def fmt(v):
            return f"{v:.1f}{unit}" if v is not None else "—"

        if invert:
            strong = f"< {fmt(abs(good_t)) if good_t is not None else '—'}"
            fair = (
                f"{fmt(abs(warn_t))} – {fmt(abs(good_t))}"
                if good_t is not None and warn_t is not None else "—"
            )
            weak = f"> {fmt(abs(warn_t)) if warn_t is not None else '—'}"
        else:
            strong = f"> {fmt(good_t)}"
            fair = (
                f"{fmt(warn_t)} – {fmt(good_t)}"
                if good_t is not None and warn_t is not None else "—"
            )
            weak = f"< {fmt(warn_t)}"

        _write_data_cell(sheet, row, 1, metric_label, ti)
        _write_data_cell(sheet, row, 2, strong, ti)
        _write_data_cell(sheet, row, 3, fair, ti)
        _write_data_cell(sheet, row, 4, weak, ti)
        _write_data_cell(sheet, row, 5, source, ti)
        desc_cell = _write_data_cell(sheet, row, 6, description, ti)
        # Single-line rendering: disable wrap so the column width controls
        # the display. Rows keep the default height.
        desc_cell.alignment = px_align(h='left', wrap=False)
        row += 1

    row += 2
    _write_footer(sheet, row, 1)


def _write_analysis(workbook, sheet, metrics: PortfolioMetrics):
    """Multi-Purpose Analysis: contribution bars, breakdowns, rebalancing actions."""
    _apply_title(sheet, 1, 1, "Multi-Purpose Analysis")
    row = 3

    df = metrics.holdings_df
    if df.empty:
        _write_data_cell(sheet, row, 1, "No data available", 0)
        return

    # Return contribution by holding
    _apply_subtitle(sheet, row, 1, "Return Contribution by Holding")
    row += 1
    for c, h in enumerate(["Name", "ISIN", "Ticker", "Weight %", "Gain %", "Contribution"]):
        _apply_header(sheet, row, c + 1, h)
    row += 1
    start_row = row
    contrib_rows = []
    for _, r in df.iterrows():
        contrib = r.get("weight_pct", 0) * r.get("gain_pct", 0) / 100
        contrib_rows.append((r, contrib))
    contrib_rows.sort(key=lambda x: -x[1])
    for ti, (r, contrib) in enumerate(contrib_rows):
        _write_data_cell(sheet, row, 1, r.get("name", ""), ti)
        _write_data_cell(sheet, row, 2, r.get("isin", ""), ti)
        _write_data_cell(sheet, row, 3, r.get("ticker", ""), ti)
        _write_data_cell(sheet, row, 4, r.get("weight_pct", 0), ti, is_number=True)
        _write_data_cell(sheet, row, 5, r.get("gain_pct", 0), ti, is_number=True)
        _write_data_cell(sheet, row, 6, contrib, ti, is_number=True)
        row += 1

    chart = _make_bar("Return Contribution by Holding", sheet,
                      1, [(6, "Return Contribution")], start_row, row - 1, width=20, height=12)
    # CHARTS DISABLED: sheet.add_chart(chart, "H3")

    row += 1

    # Breakdown by asset class
    _apply_subtitle(sheet, row, 1, "Breakdown by Asset Class")
    row += 1
    class_group = df.groupby("asset_class").agg(
        total_value=("current_value", "sum"),
        avg_gain=("gain_pct", "mean"),
        count=("ticker", "count"),
    ).reset_index().sort_values("avg_gain", ascending=False)
    for c, h in enumerate(["Asset Class", "Total Value", "Avg Gain %", "# Holdings"]):
        _apply_header(sheet, row, c + 1, h)
    row += 1
    for ti, (_, r) in enumerate(class_group.iterrows()):
        _write_data_cell(sheet, row, 1, r["asset_class"], ti, asset_class=r["asset_class"])
        _write_data_cell(sheet, row, 2, r["total_value"], ti, is_number=True, num_fmt='#,##0.00')
        _write_data_cell(sheet, row, 3, r["avg_gain"], ti, is_number=True)
        _write_data_cell(sheet, row, 4, r["count"], ti)
        row += 1

    row += 1

    # Breakdown by geography (equity only)
    _apply_subtitle(sheet, row, 1, "Breakdown by Geography (Equity Only)")
    row += 1
    if not metrics.allocation_by_geo.empty:
        geo_gains: dict[str, list[float]] = {}
        for _, r in df.iterrows():
            if r.get("asset_class") != "Equities":
                continue
            geo_str = r.get("geography", "")
            gain = r.get("gain_pct", 0)
            if "," in str(geo_str):
                for part in str(geo_str).split(","):
                    part = part.strip()
                    if ":" in part:
                        geo_name = part.split(":")[0].strip()
                        geo_gains.setdefault(geo_name, []).append(gain)
            else:
                geo_gains.setdefault(str(geo_str), []).append(gain)

        for c, h in enumerate(["Geography", "Weight % (within Equity)", "Avg Gain %"]):
            _apply_header(sheet, row, c + 1, h)
        row += 1
        for ti, (_, r) in enumerate(metrics.allocation_by_geo.iterrows()):
            cat = r["category"]
            gains = geo_gains.get(cat, [])
            avg_gain = sum(gains) / len(gains) if gains else 0
            _write_data_cell(sheet, row, 1, cat, ti, geography=cat)
            _write_data_cell(sheet, row, 2, r["weight_pct"], ti, is_number=True)
            _write_data_cell(sheet, row, 3, avg_gain, ti, is_number=True)
            row += 1
    else:
        _write_data_cell(sheet, row, 1, "No equity holdings", 0)
        row += 1




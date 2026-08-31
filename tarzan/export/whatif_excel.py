"""Excel export for the ad-hoc what-if comparison (``scripts/whatif.py``).

Single sheet, **portfolios as columns**: a specs block (summary + per-
instrument weights) followed by matrices for funded-capital allocation,
notional exposure, equity geography, and returns & risk. Any number of
portfolios (the real "Current" one plus each column of the weights CSV)
render side by side, so adding a portfolio just adds a column.

Colors come from the shared taxonomy in ``tarzan.export._format``.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tarzan.export._format import (
    asset_class_color,
    eur_smart,
    geo_color,
    short_instrument_name,
)

logger = logging.getLogger(__name__)

_C = {
    "header": "1E293B",
    "band": "5B5BD6",
    "alt": "F8FAFF",
    "white": "FFFFFF",
    "text": "1E293B",
    "muted": "64748B",
    "green": "16A34A",
    "red": "DC2626",
    "amber": "D97706",
    "border": "CBD5E1",
}

# Categorical series palette in a FIXED order, never cycled for identity, and
# the dash patterns that pair with it. Validated in OKLab (ΔE ×100) with
# protan/deutan/tritan simulation: adjacent pairs in this order score 19.6 for
# normal vision (pass, floor 15) and 6.1 at worst under CVD (tritan, in the 6-8
# band) — which is legal ONLY alongside a secondary encoding, hence the dashes.
# Across ALL pairs, not just adjacent ones, the worst separation collapses to 7.1
# normal / 3.9 CVD: eight hues cannot give 30+ series individual identity, so
# above ~8 series colour is a grouping cue and the ranked legend is what carries
# lookup. Identity is assigned from a portfolio's fixed position, never from its
# rank in a panel, so a chart with fewer series never repaints the survivors.
_SERIES_HUES = ["2A78D6", "EB6834", "1BAF7A", "EDA100",
                "E87BA4", "008300", "4A3AA7", "E34948"]
_SERIES_DASHES = ["-", (0, (5, 1.5)), (0, (1.5, 1.5)), (0, (6, 1.5, 1.5, 1.5))]

from tarzan.models.taxonomy import ORDER_WHATIF as _ORDER_WHATIF, GEO_ORDER as _GEO_REG

_ASSET_ORDER = list(_ORDER_WHATIF)
# The what-if workbook also renders an explicit "Other" geo bucket at the end.
_GEO_ORDER = list(_GEO_REG) + ["Other"]

_PCOL0 = 3  # portfolios start at column C (A = ticker/category, B = description)


def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)


def _font(size=10, bold=False, color="1E293B", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)


def _border():
    s = Side(style="thin", color=_C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _dev_color(delta, tol):
    if delta is None or delta != delta or tol <= 0:
        return _C["text"]
    a = abs(delta)
    if a <= tol:
        return _C["green"]
    if a <= 2 * tol:
        return _C["amber"]
    return _C["red"]


# ---------------------------------------------------------------------------
# Building blocks
# ---------------------------------------------------------------------------

# Per-portfolio lineage notes, kept IN CODE (no external input file) so the
# workbook is self-contained. Keyed by portfolio name; unknown names → no note.
_PORTFOLIO_NOTES: dict[str, str] = {
    "start_cl2": "Starting portfolio (ex current_tgt/v11): NTSG efficient-core + CL2 (2x USA) + value/momentum tilt (XDEV/XDEM) + long-duration barbell X25E. ~1.28x, geography on target.",
    "fac_growth": "Equity-forward integrated factor tilt (Avantis AVWC+AVWS + XDEM momentum): value+size+profitability+momentum. Higher return, Sharpe ~0.49.",
    "rb_nodur": "Risk-balanced WITHOUT long duration (no X25E): NTSG 38, gold 18, trend 17, carry 8, AVWS 8. Best Sharpe/drawdown of the legacy set; the high carry inflates its Sharpe a bit.",
    "ref_gold22": "Max-Sharpe reference (gold 22%, carry 10%): NTSG + SC2X 2x + Avantis value + trend/gold/carry/long-duration. Best risk-adjusted but gold weight high (community 7-14%).",
    "base_xmme": "Structural base (carry 5%, XMME): SC2X capped 10%, leverage from NTSG (90/60), gold 15%, X25E 4%. Global ~62% USA, ~1.26x.",
    "base_avem": "base_xmme with XMME->AVEM (EM value+profitability on the FF-Emerging legs, +~EUR10k/15y). Keeps X25E 4%.",
    "base_mom": "base + XDEM momentum (value+momentum barbell), carry 5%, X25E 4%. Momentum a slight in-window drag.",
    "ntsg30_dur": "NTSG 30 (carry 5, X25E 5), factors trimmed. Tests more balanced core vs factor tilt.",
    "ntsg35_dur": "NTSG 35 (carry 5, X25E 5), factors trimmed further. More core, fewer factors.",
    "nodur_gold": "base_avem WITHOUT X25E, freed capital to gold+trend (gold 17, trend 18). Best drawdown of the no-duration set; +CAGR vs base_avem, Sharpe 0.50.",
    "target_fac": "TARGET #1 (return-forward): no X25E, gold/trend capped 15, freed capital to the FACTORS (AVWC 10 / AVWS 9 / AVEM 4). Highest CAGR (8.1%, EUR335k/15y), Sharpe 0.50, drawdown -30.6%.",
    "target_mix": "TARGET #2 (balanced): no X25E, gold/trend capped 15, freed to NTSG+factors (NTSG 35, AVWC 9/AVWS 8/AVEM 3). CAGR 8.06%, Sharpe 0.49, drawdown -30.2%.",
    "nodur_ntsg": "no X25E, gold/trend 15, freed capital to NTSG (37). More core: good return but worse Sharpe/drawdown than the factor route.",
    "mom_light": "target_fac + a SMALL momentum sleeve (XDEM 4) funded by trimming the mildest value leg (AVWC 10 to 6); momentum is 17% of the factor sleeve. Tests the marginal effect of pairing momentum with value at low conviction.",
    "mom_6040": "target_fac rebuilt at the classic 60/40 VALUE/MOMENTUM split of the factor sleeve (value 14 = AVWC 5 / AVWS 6 / AVEM 3, momentum 9 = XDEM). Rationale is diversification, not return: a value tilt implicitly shorts momentum, and the pairing is worth holding even at zero momentum premium. Return-neutral to slightly negative, best drawdown of the momentum set.",
    "target_fac_cl2": "target_fac with the levered sleeve SWAPPED from SC2X to CL2. Motivation: SC2X launched in July 2026 (19 days of real history, TER 1.40%), CL2 has ~5 years and costs 0.50%. The trade-off is geographic, not statistical: CL2 is 2x MSCI USA (100% US) where SC2X is 2x ACWI, so the whole leveraged sleeve concentrates in the US.",
    "target_mix_cl2": "target_mix (NTSG 35) with the same SC2X to CL2 swap: cheaper and far less synthetic, at the cost of a US-only leveraged sleeve.",
    "mom_6040_cl2": "mom_6040 (60/40 value/momentum sleeve) with the same SC2X to CL2 swap - the momentum variant on a cheaper, longer-lived leverage vehicle.",
    "bench_eq100": "Benchmark: 100% world equity (MSCI World). Equity-risk reference: low Sharpe, drawdown -52%.",
    "bench_6040": "Benchmark 60/40: 60% MSCI World + 40% aggregate bond (EUR hedged).",
    "bench_golden_butterfly": "Benchmark Golden Butterfly (global adaptation): 20 world / 20 small-value / 20 long-govt / 20 cash / 20 gold.",
    "bench_all_weather": "Benchmark All Weather / All Seasons: 30 equity / 40 long-govt / 15 aggregate / 7.5 gold / 7.5 commodity. Hurt by long bonds (2022).",
    "bench_permanent": "Benchmark Permanent Portfolio (Browne): 25 equity / 25 long-govt / 25 gold / 25 cash. Lowest drawdown, low return.",
    "bench_larry": "Benchmark Larry Portfolio (Swedroe): 30 small-value + 70 aggregate bond. Little but high-returning equity, lots of bonds.",
    "bench_golden_ratio": "Benchmark Golden Ratio (PortfolioCharts): 42 world / 26 long-govt / 16 gold / 12 small-value / 4 cash.",
    "bench_pinwheel": "Benchmark Pinwheel: ~45 equity (world+SCV+EM) / 15 bond / 15 gold / 10 cash (REIT approximated as equity).",
    "bench_swensen": "Benchmark Yale/Swensen (approx): ~65 equity (US+intl+REIT folded) / 5 EM / 30 bond (treasury+TIPS folded). REIT->equity, TIPS->bond.",
    "bench_ivy": "Benchmark Ivy (Faber 5-asset): 60 equity (US+intl+REIT folded) / 20 bond / 20 commodity.",
    "bench_cockroach": "Benchmark Cockroach (Mutiny): 25 equity / 25 long-govt / 25 gold / 25 trend. Very low drawdown thanks to trend; a minimal version of the risk-balanced thesis.",
    "bench_weird": "Benchmark Weird Portfolio: 40 world / 20 small-value / 20 long-govt / 20 gold (REIT folded to equity). Best UNLEVERED Sharpe (0.54).",
}


def _load_portfolio_notes() -> dict[str, str]:
    """Per-portfolio lineage notes (name → description), held in code (see
    :data:`_PORTFOLIO_NOTES`). No external file."""
    return dict(_PORTFOLIO_NOTES)


def _lineage_block(ws, row, portfolios, ncol):
    """Render 'How we got here' — one row per portfolio with its lineage note
    (from input/portfolio_notes.csv), so the reader can tell the candidates
    apart. Skipped entirely when no notes file is present."""
    notes = _load_portfolio_notes()
    have = [(p.name, notes.get(p.name, "")) for p in portfolios]
    if not any(desc for _, desc in have):
        return row
    row = _section_header(ws, row, "How we got here — portfolio lineage", ncol)
    for k, (name, desc) in enumerate(have):
        bg = _C["alt"] if k % 2 else _C["white"]
        a = ws.cell(row=row, column=1, value=name.replace("_", " "))
        a.font = _font(9, bold=True, color=_C["band"]); a.fill = _fill(bg)
        a.border = _border(); a.alignment = _align("left", "top")
        b = ws.cell(row=row, column=2, value=desc)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncol)
        b.font = _font(9, color=_C["muted"]); b.fill = _fill(bg); b.border = _border()
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 42
        row += 1
    return row + 1


def _section_header(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(11, bold=True, color=_C["white"])
    c.fill = _fill(_C["band"])
    return row + 1


def _merge_label(ws, row, text, *, bold=True, color=None, bg=None, italic=False):
    """Write a row label merged across columns A:B (so tables whose data
    starts at column C don't show an empty B)."""
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        if bg:
            c.fill = _fill(bg)
            c.border = _border()
    a = ws.cell(row=row, column=1, value=text)
    a.font = _font(9, bold=bold, color=color or _C["text"], italic=italic)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    return a


def _col_headers(ws, row, portfolios, label_hdr, with_target, second_hdr=None):
    """Write the label (+ optional second label) + one-column-per-portfolio
    (+ optional Target) header."""
    h = ws.cell(row=row, column=1, value=label_hdr)
    h.font = _font(9, bold=True, color=_C["white"])
    h.fill = _fill(_C["header"])
    h.border = _border()
    if second_hdr is not None:
        h2 = ws.cell(row=row, column=2, value=second_hdr)
        h2.font = _font(9, bold=True, color=_C["white"])
        h2.fill = _fill(_C["header"])
        h2.border = _border()
    else:
        # No second column in use → merge A:B so B isn't an empty column.
        c2 = ws.cell(row=row, column=2)
        c2.fill = _fill(_C["header"])
        c2.border = _border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    for j, p in enumerate(portfolios):
        # Display name with underscores as spaces so it wraps cleanly at word
        # boundaries (not mid-word) in the narrow per-portfolio columns.
        c = ws.cell(row=row, column=_PCOL0 + j, value=p.name.replace("_", " "))
        c.font = _font(9, bold=True, color=_C["white"])
        c.fill = _fill(_C["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 42
    tcol = None
    if with_target:
        tcol = _PCOL0 + len(portfolios)
        c = ws.cell(row=row, column=tcol, value="Target")
        c.font = _font(9, bold=True, color=_C["white"])
        c.fill = _fill(_C["header"])
        c.alignment = _align("center")
        c.border = _border()
    return tcol


def _alloc_matrix(ws, row, title, labels, portfolios, value_of, *,
                  target=None, tolerance=1.5, swatch=None) -> dict:
    """Allocation/geo matrix: rows = labels, cols = portfolios (+ Target).

    ``value_of(p, label)`` returns a percentage (0..100+). Portfolio cells
    are traffic-light coloured by their deviation from target when a target
    is supplied.
    """
    ncols = _PCOL0 + len(portfolios) + (1 if target is not None else 0) - 1
    row = _section_header(ws, row, title, ncols)
    tcol = _col_headers(ws, row, portfolios, "Category", target is not None)
    header_row = row
    row += 1
    first = row
    for i, lbl in enumerate(labels):
        bg = _C["alt"] if i % 2 else _C["white"]
        _merge_label(ws, row, lbl, bold=True,
                     color=(swatch(lbl) if swatch else _C["text"]), bg=bg)
        t = target.get(lbl, 0.0) if target is not None else None
        for j, p in enumerate(portfolios):
            v = value_of(p, lbl)
            c = ws.cell(row=row, column=_PCOL0 + j, value=v / 100.0)
            c.number_format = "0.0%"
            c.alignment = _align("center")
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _font(9, color=(_dev_color(v - t, tolerance) if t is not None else _C["text"]))
        if tcol is not None:
            c = ws.cell(row=row, column=tcol, value=t / 100.0)
            c.number_format = "0.0%"
            c.alignment = _align("center")
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _font(9, bold=True, color=_C["muted"])
        row += 1
    return {"header_row": header_row, "first": first, "last": row - 1, "tcol": tcol}


def _plain_row(ws, row, label, portfolios, text_of, color=None):
    """A summary row (label + one text value per portfolio), no target."""
    _merge_label(ws, row, label, bold=True, color=color)
    for j, p in enumerate(portfolios):
        c = ws.cell(row=row, column=_PCOL0 + j, value=text_of(p))
        c.alignment = _align("center")
        c.font = _font(9, color=color or _C["text"])
    return row + 1


def _instrument_matrix_block(ws, row, portfolios) -> int:
    """Per-instrument weight matrix, grouped by asset class + role (sub-class) —
    the same layout as the newsletter's "Instruments × portfolios" block."""
    ncols = _PCOL0 + len(portfolios) - 1
    row = _section_header(ws, row, "Instruments \u00d7 portfolios", ncols)
    _col_headers(ws, row, portfolios, "Ticker", with_target=False, second_hdr="Description")
    row += 1

    reps: dict = {}
    for p in portfolios:
        for it in p.items:
            reps.setdefault(it.bare, it)

    def _cls(it):
        return (it.holding.asset_class.value
                if getattr(it.holding, "asset_class", None) else "Other")

    def _role(it):
        return getattr(it.holding, "role", "") or "\u2014"

    by_cls: dict = {}
    for it in reps.values():
        by_cls.setdefault(_cls(it), []).append(it)
    ordered = [c for c in _ASSET_ORDER if c in by_cls] + \
              [c for c in by_cls if c not in _ASSET_ORDER]
    name_by = {b: short_instrument_name(it.holding.name or b, 46)
               for b, it in reps.items()}
    wmaps = [p.weights() for p in portfolios]

    i = 0
    for cls in ordered:
        color = asset_class_color(cls)
        # Class header spanning the full width, coloured swatch via font colour.
        ch = _merge_label(ws, row, cls, bold=True, color=color, bg=_C["alt"])
        for j in range(len(portfolios)):
            cc = ws.cell(row=row, column=_PCOL0 + j)
            cc.fill = _fill(_C["alt"])
            cc.border = _border()
        row += 1
        roles: dict = {}
        for it in by_cls[cls]:
            roles.setdefault(_role(it), []).append(it)
        for role, its in roles.items():
            if role and role != "\u2014":
                r = ws.cell(row=row, column=1, value=f"   {role}")
                r.font = _font(8, italic=True, color=_C["muted"])
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                row += 1
            for it in sorted(its, key=lambda x: x.bare):
                bg = _C["alt"] if i % 2 else _C["white"]
                i += 1
                a = ws.cell(row=row, column=1, value=it.bare)
                a.font = _font(9, bold=True)
                a.fill = _fill(bg)
                a.border = _border()
                b = ws.cell(row=row, column=2, value=name_by.get(it.bare, ""))
                b.font = _font(9, color=_C["muted"])
                b.fill = _fill(bg)
                b.border = _border()
                for j, wm in enumerate(wmaps):
                    c = ws.cell(row=row, column=_PCOL0 + j,
                                value=(wm[it.bare] / 100.0 if it.bare in wm else None))
                    c.number_format = "0.0%"
                    c.alignment = _align("center")
                    c.fill = _fill(bg)
                    c.border = _border()
                row += 1
    return row + 1


# Same metric set as the newsletter "Portfolio risk metrics" block (incl. Ulcer).
_METRIC_ROWS = [
    ("CAGR", "cagr", "%"), ("Volatility (ann.)", "volatility", "%"),
    ("Sharpe", "sharpe", ""), ("Sortino", "sortino", ""),
    ("Max Drawdown", "max_drawdown", "%"), ("Ulcer index", "ulcer_index", "%"),
    ("VaR 95% (daily)", "var_95", "%"), ("CVaR 95% (daily)", "cvar_95", "%"),
    ("Beta vs S&P 500", "beta", ""), ("Alpha (ann.)", "alpha", "%"),
]


def _inject_ulcer(portfolios) -> None:
    """Ensure each portfolio's aligned metrics carry an Ulcer index (computed
    from its NAV, currency-independent) — mirrors the newsletter block."""
    from tarzan.engine.stats import compute_ulcer_index
    for p in portfolios:
        nav = getattr(p, "nav", None)
        if nav is None:
            continue
        for attr in ("metrics_aligned_eur", "metrics_aligned_usd"):
            m = getattr(p, attr, None)
            if isinstance(m, dict) and "ulcer_index" not in m:
                try:
                    m["ulcer_index"] = compute_ulcer_index(nav)
                except Exception:  # noqa: BLE001
                    m["ulcer_index"] = None


def _metric_value(metrics, key, unit):
    v = (metrics or {}).get(key)
    if v is None or (isinstance(v, float) and v != v):
        return "n/a"
    return f"{v:.2f}{unit}"


def _metrics_block(ws, row, portfolios, attr, ccy_label) -> int:
    """Render one currency block of the metrics matrix (EUR or USD)."""
    ncols = _PCOL0 + len(portfolios) - 1
    rf = next(((getattr(p, attr, {}) or {}).get("risk_free")
               for p in portfolios if getattr(p, attr, None)), None)
    rf_s = f"  ·  risk-free {rf:.2f}%" if isinstance(rf, (int, float)) else ""
    _merge_label(ws, row, f"{ccy_label}{rf_s}", bold=True, color=_C["band"])
    row += 1
    for i, (label, key, unit) in enumerate(_METRIC_ROWS):
        bg = _C["alt"] if i % 2 else _C["white"]
        _merge_label(ws, row, label, bold=True, bg=bg)
        for j, p in enumerate(portfolios):
            c = ws.cell(row=row, column=_PCOL0 + j,
                        value=_metric_value(getattr(p, attr, {}), key, unit))
            c.alignment = _align("center")
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _font(9)
        row += 1
    # Two derived, money-terms rows over the FULL aligned window: cumulative total
    # return, and the future value of €100k invested at t0 — the same compounded
    # CAGR, expressed as growth and as euros.
    def _growth(p):
        m = getattr(p, attr, {}) or {}
        cagr, w = m.get("cagr"), getattr(p, "window", None)
        if cagr is None or (isinstance(cagr, float) and cagr != cagr) or not w:
            return None
        yrs = (w[1] - w[0]).days / 365.25
        return (1.0 + cagr / 100.0) ** yrs        # final / initial multiple
    for off, (label, fmt) in enumerate((("Total return (cumulative)", "pct"),
                                        ("€100k at t0 →", "eur"))):
        bg = _C["alt"] if (len(_METRIC_ROWS) + off) % 2 else _C["white"]
        _merge_label(ws, row, label, bold=True, bg=bg)
        for j, p in enumerate(portfolios):
            g = _growth(p)
            txt = ("n/a" if g is None else
                   f"{(g - 1) * 100:.0f}%" if fmt == "pct" else eur_smart(100_000 * g))
            c = ws.cell(row=row, column=_PCOL0 + j, value=txt)
            c.alignment = _align("center"); c.fill = _fill(bg)
            c.border = _border(); c.font = _font(9)
        row += 1
    return row


def _risk_matrix(ws, row, portfolios) -> int:
    _inject_ulcer(portfolios)
    ncols = _PCOL0 + len(portfolios) - 1
    w = next((p.window for p in portfolios if getattr(p, "window", None)), None)
    win = f"{w[0]:%Y-%m} → {w[1]:%Y-%m}" if w else "aligned"
    row = _section_header(ws, row, f"Portfolio metrics — single aligned history ({win})", ncols)
    _col_headers(ws, row, portfolios, "Metric", with_target=False)
    row += 1
    row = _metrics_block(ws, row, portfolios, "metrics_aligned_eur", "EUR numeraire (unhedged)")
    row += 1
    row = _metrics_block(ws, row, portfolios, "metrics_aligned_usd", "USD numeraire")
    return row + 1


# The portfolios under active consideration: pinned to the top of the Summary,
# given a Monte-Carlo fan chart and drawn heavy on the growth chart. Everything
# else is ranked below them. target_fac_cl2 is the one currently written into
# input/targets_per_holding.csv (the live rebalancing target); the others stay
# starred because they remain in the running.
# A TUPLE, not a set: the order is meaningful. The first entry is the allocation
# currently written into input/targets_per_holding.csv, and the Method sheet's
# worked example follows this order so it documents the live target rather than
# whichever name a set happened to yield first.
_TARGETS = ("target_fac_cl2", "target_fac", "target_mix", "mom_6040")

# Windows for the summary (label, start-date); FULL first, then restricted.
_SUMMARY_WINDOWS = (("FULL 2000-26", "2000-08-31"), ("2011-26", "2011-01-01"),
                    ("2020-26", "2020-01-01"), ("2026 YTD", "2026-01-01"))
# Per-window KPIs (label, full_metrics key or 'calmar', number format).
_WIN_METRICS = (("CAGR", "cagr", "0.0"), ("Vol%", "volatility", "0.0"),
                ("Shrp", "sharpe", "0.00"), ("Sort", "sortino", "0.00"),
                ("MaxDD%", "max_drawdown", "0.0"), ("Calmr", "calmar", "0.00"),
                ("β", "beta", "0.00"), ("α%", "alpha", "0.0"))


def _clean_num(v):
    """None for missing/NaN so the cell shows n/a instead of a broken value."""
    return None if (v is None or (isinstance(v, float) and v != v)) else v


def _mc_fan_png(nav, name, *, years=15, n_sims=1000, block=1, anchor=100_000):
    """Monte-Carlo fan chart (block-bootstrap) of €``anchor`` over ``years``: the
    5/25/50/75/95th-percentile cone of terminal wealth across ``n_sims`` reshuffled
    paths of the portfolio's own MONTHLY returns — the same frequency and block
    size as the MC15y KPI (:func:`tarzan.engine.robustness.block_bootstrap`), so
    the picture and the column cannot tell different stories. ``block`` is in
    MONTHS. Returns a PNG BytesIO (or None if the history is too short)."""
    try:
        import numpy as np
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from io import BytesIO
    except Exception:  # noqa: BLE001
        return None
    if nav is None or getattr(nav, "empty", True):
        return None
    r_d = nav.pct_change().replace([np.inf, -np.inf], np.nan).dropna()
    r = ((1.0 + r_d).resample("ME").prod() - 1.0).dropna().values
    months, n = int(years * 12), len(r)
    if n < max(block * 3, 24):
        return None
    nb = int(np.ceil(months / block))
    rng = np.random.default_rng(12345)                 # fixed seed → reproducible
    paths = np.empty((n_sims, months))
    for i in range(n_sims):
        starts = rng.integers(0, max(1, n - block), size=nb)
        seq = np.concatenate([r[s:s + block] for s in starts])[:months]
        paths[i] = anchor * np.cumprod(1.0 + seq)
    t = np.arange(1, months + 1) / 12.0
    pc = {q: np.percentile(paths, q, axis=0) for q in (5, 25, 50, 75, 95)}
    fig, ax = plt.subplots(figsize=(6.2, 3.5), dpi=130)
    fig.subplots_adjust(left=0.13, right=0.99, top=0.86, bottom=0.13)
    ax.fill_between(t, pc[5], pc[95], color="#93C5FD", alpha=0.35, lw=0, label="5–95%")
    ax.fill_between(t, pc[25], pc[75], color="#3B82F6", alpha=0.40, lw=0, label="25–75%")
    ax.plot(t, pc[50], color="#1E3A8A", lw=2.0, zorder=3, label="median")
    ax.axhline(anchor, color="#94A3B8", lw=0.8, ls=":")
    _ink, _mut, _bnd = f"#{_C['text']}", f"#{_C['muted']}", f"#{_C['band']}"
    ax.set_yscale("log")
    ax.set_title(f"Monte-Carlo {years}y — {name}  (€{anchor // 1000}k, monthly block-bootstrap)",
                 fontsize=10, fontweight="bold", color=_ink, loc="left")
    ax.set_xlabel("years", fontsize=8, color=_mut)
    ax.grid(alpha=0.3, which="both")
    for q in (95, 50, 5):
        ax.annotate(f" {q}%: €{pc[q][-1] / 1000:.0f}k", xy=(t[-1], pc[q][-1]),
                    xytext=(3, 0), textcoords="offset points", fontsize=7.5,
                    va="center", ha="left", color=_bnd, annotation_clip=False)
    ax.set_xlim(0, years + 2.2)
    ax.legend(fontsize=7, loc="upper left", frameon=False)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"€{v / 1000:.0f}k"))
    ax.tick_params(labelsize=7.5, colors=_mut)
    buf = BytesIO()
    fig.savefig(buf, format="png"); plt.close(fig); buf.seek(0)
    return buf


def _growth_windows_png(portfolios, *, anchor=100.0, metrics_of=None):
    """Growth-of-100 (log) for ALL portfolios, one panel per window.

    Each panel carries its OWN legend, ranked by that window's cumulative return
    (best first) and annotated with the window's ``(CAGR, Vol, Sharpe)`` — with
    30+ series a line cannot be traced back to a name by colour alone, so the
    ranked legend is what makes the chart answerable ("who is on top, at what
    risk"). Series identity is hue × dash from a portfolio's FIXED position (see
    :data:`_SERIES_HUES`), never from its rank, so it is stable across panels.
    Targets are black and heavy; benchmarks are lightened as context.

    ``metrics_of(nav, start) -> dict`` supplies the per-window metrics; passing
    the Summary's own accessor keeps legend and table on one source of truth.
    Returns a PNG BytesIO (or None)."""
    try:
        import numpy as np
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.ticker import FixedLocator, NullLocator
        from io import BytesIO
    except Exception:  # noqa: BLE001
        return None
    navs = {p.name: p.synth_nav for p in portfolios
            if getattr(p, "synth_nav", None) is not None and not p.synth_nav.empty}
    if not navs:
        return None
    names = list(navs)
    targets = _TARGETS
    # Stable identity: hue varies fastest so neighbours in the fixed order get the
    # validated adjacent-pair separation; the dash changes every 8 series.
    plain = [n for n in names if n not in targets]
    style = {n: (f"#{_SERIES_HUES[i % len(_SERIES_HUES)]}",
                 _SERIES_DASHES[(i // len(_SERIES_HUES)) % len(_SERIES_DASHES)])
             for i, n in enumerate(plain)}
    for i, n in enumerate(n for n in names if n in targets):   # targets: black, distinct dash
        style[n] = ("#000000", _SERIES_DASHES[i % len(_SERIES_DASHES)])
    wins = (("2000-08-31", "FULL 2000-26"), ("2011-01-01", "2011-26"),
            ("2020-01-01", "2020-26"), ("2026-01-01", "2026 YTD"))
    fig, axes = plt.subplots(4, 1, figsize=(9.6, 21.0), dpi=120,
                             gridspec_kw={"hspace": 0.62})
    for ax, (start, title) in zip(axes, wins):
        sub = {n: navs[n].loc[start:] for n in names if not navs[n].loc[start:].empty}
        if not sub:
            continue
        cs = max(s.index.min() for s in sub.values())
        lo_v, hi_v = 1e18, -1e18
        entries = []                                  # (final growth, handle, label)
        for n, s in sub.items():
            s = s.loc[cs:]
            if s.empty or s.iloc[0] <= 0:
                continue
            g = s / s.iloc[0] * anchor
            lo_v = min(lo_v, float(g.min())); hi_v = max(hi_v, float(g.max()))
            is_t, is_b = n in targets, n.startswith("bench_")
            col, dash = style[n]
            ln, = ax.plot(g.index, g.values, color=col, ls=dash,
                          lw=(2.6 if is_t else 1.0),
                          alpha=(1.0 if is_t else 0.55 if is_b else 0.9),
                          zorder=(4 if is_t else 2))
            m = metrics_of(navs[n], start) if metrics_of is not None else {}
            cg, vol, shp = m.get("cagr"), m.get("volatility"), m.get("sharpe")
            stats = (f" ({cg:.1f}%, {vol:.1f}%, {shp:.2f})"
                     if None not in (cg, vol, shp) else "")
            label = ("★ " if is_t else "") + n.replace("_", " ") + stats
            entries.append((float(g.iloc[-1]), ln, label))
        # Legend ranked by this window's cumulative return, best first.
        entries.sort(key=lambda e: -e[0])
        ax.legend([e[1] for e in entries], [e[2] for e in entries],
                  fontsize=5.0, ncol=3, loc="upper left", bbox_to_anchor=(0, -0.09),
                  frameon=False, borderaxespad=0, handlelength=3.2, columnspacing=1.2,
                  labelspacing=0.28,
                  title="ranked by return · (CAGR, Vol, Sharpe) · ★ target",
                  title_fontsize=6.5, alignment="left")
        ax.set_yscale("log")
        ax.set_title(f"Cumulative return — {title} (log)  ·  heavy black = target, faded = benchmark",
                     fontsize=10, fontweight="bold", color=f"#{_C['text']}", loc="left")
        ax.set_ylabel("cumulative return", fontsize=8, color=f"#{_C['muted']}")
        ax.grid(alpha=0.3, which="both")
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)
        ax.tick_params(labelsize=7.5, colors=f"#{_C['muted']}")
        # Log axis, but with ticks fixed at sensible growth levels and labelled as
        # cumulative % return (index − 100), so it reads "+100%" not "10^2".
        if hi_v / max(lo_v, 1.0) > 3.0:          # wide window → coarse tick levels
            lv = [40, 60, 80, 100, 150, 200, 300, 500, 800, 1200, 2000, 3000, 5000]
        else:                                     # narrow window (e.g. 2026 YTD) → fine
            lv = [70, 80, 85, 90, 95, 100, 105, 110, 115, 120, 130, 140, 150]
        ax.yaxis.set_major_locator(FixedLocator(lv))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v - anchor:+.0f}%"))
        ax.yaxis.set_minor_locator(NullLocator())
    axes[-1].set_xlabel("year", fontsize=8, color=f"#{_C['muted']}")
    buf = BytesIO()
    # bbox_inches="tight" so the per-panel legends are never clipped; the caller
    # derives the Excel row span from the resulting image height.
    fig.savefig(buf, format="png", bbox_inches="tight"); plt.close(fig); buf.seek(0)
    return buf


def _summary_sheet(wb, portfolios) -> None:
    """First sheet: one ROW per portfolio, ordered as a decision table.

    Columns lead with what an allocation decision actually turns on — the 15y
    Monte-Carlo projection (median / bad case / dispersion), the €100k outcome
    and its multiplier, the rolling 5/10/15y holds, then the notional asset split
    and equity geography — followed by leverage and only THEN the per-window KPI
    detail (CAGR/Vol/Sharpe/Sortino/MaxDD/Calmar/β/α × 4 windows, each window
    boxed by vertical rules). The instrument recap stays last: it is a wide text
    column and would push the numbers off-screen anywhere else.

    Rows put the TARGETS on top, then rank the rest by MC 15y median descending
    (best expected outcome first), tie-broken by MC band ascending (tighter
    dispersion first — for the band, lower is better)."""
    from tarzan.engine import robustness as _rob
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["band"]
    targets = _TARGETS

    # Global equity benchmark for beta/alpha = MSCI ACWI, built as an ACWI-weighted
    # blend of the geo proxies (all span ~2000-2026, unlike the single VTWSX/ACWI
    # proxy which is stale post-2019). Falls back to bench_eq100 (MSCI World).
    bench_nav = None
    try:
        import pandas as _pdp
        from tarzan.data import proxy_data as _pd
        from tarzan.engine import synthetic as _syn
        _W = {"USA": 0.62, "Dev ex-USA ex-EMU ex-JP": 0.14, "Eurozone EMU": 0.09,
              "Japan": 0.05, "Emerging Markets": 0.10}
        _pr, _ = _pd.proxy_returns_for(set(_W))
        _cols = {k: _pr[k] for k in _W if _pr.get(k) is not None and not _pr[k].empty}
        if _cols:
            _df = _pdp.DataFrame(_cols).dropna()
            _w = _pdp.Series({k: _W[k] for k in _cols}); _w = _w / _w.sum()
            bench_nav = _syn.returns_to_price((_df * _w).sum(axis=1))
    except Exception:  # noqa: BLE001
        bench_nav = None
    if bench_nav is None:
        _bp = next((q for q in portfolios if q.name == "bench_eq100"), None)
        bench_nav = getattr(_bp, "synth_nav", None) if _bp is not None else None

    # Time-varying risk-free (for per-window Sharpe/Sortino) over the data range.
    rf_ann = rfd = None
    try:
        from tarzan.data import proxy_data as _pd2
        _navs = [p.synth_nav for p in portfolios if getattr(p, "synth_nav", None) is not None]
        if _navs:
            _lo = min(n.index.min() for n in _navs); _hi = max(n.index.max() for n in _navs)
            rf_ann = _pd2.risk_free_annual(_lo, _hi); rfd = _pd2.risk_free_daily(_lo, _hi)
    except Exception:  # noqa: BLE001
        rf_ann = rfd = None

    def _wm(pnav, start):
        if pnav is None or getattr(pnav, "empty", True):
            return {}
        s = pnav.loc[start:]
        if len(s) < 60:
            return {}
        b = bench_nav.loc[start:] if bench_nav is not None else None
        return _rob.full_metrics(s, b, risk_free=rf_ann, rf_daily=rfd)

    # Column order, left to right: name, the decision block (MC → geography),
    # leverage, the per-window KPI detail, and the wide instrument recap last.
    fixed_pre = [("Portfolio", 19)]
    decision = [("MC15y med", 8), ("MC15y p5", 7), ("MC15y band", 9),
                ("DD str p5", 9), ("corr→1 ΔDD", 10), ("cond CAGR", 9),
                ("€100k→15y", 9),
                ("×mult", 6),
                ("R5y med", 7), ("R5y p5", 7), ("R10y med", 8), ("R10y p5", 7),
                ("R15y med", 8), ("R15y p5", 7),
                ("Eq", 5), ("Bond", 5), ("Gold", 5), ("Trend", 5), ("Comm", 5),
                ("USA", 5), ("DevexUS", 7), ("EMU", 5), ("JP", 5), ("EM", 5),
                ("Lev", 5)]
    fixed_post = [("Instruments (ticker weight)", 56)]
    widths = ([w for _, w in fixed_pre] + [w for _, w in decision]
              + [6] * (len(_WIN_METRICS) * len(_SUMMARY_WINDOWS))
              + [w for _, w in fixed_post])
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ncol = len(widths)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value="SUMMARY — all portfolios · full KPI set per window (★ = target)")
    t.font = _font(15, bold=True, color=_C["white"]); t.fill = _fill(_C["header"])
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.cell(row=2, column=1, value=(
        "Per window (FULL 2000-26, 2011-26, 2020-26, 2026-YTD): CAGR·Vol·Sharpe·Sortino·"
        "MaxDD·Calmar (EUR, Calmar=CAGR/|MaxDD|), β/α vs MSCI ACWI (α annualised, Jensen rf-tv). "
        "Vol/Sharpe/Sortino are measured on MONTHLY returns: on the reconstructed history the "
        "daily frequency carries non-synchronous pricing noise that inflated vol by ~1.4x (and so "
        "depressed Sharpe by ~30%), worse the more diversified the portfolio. MaxDD stays on the "
        "daily path (a month-end one would erase intra-month crashes) and β/α stay daily (the "
        "noise is common to portfolio and benchmark, so it cancels in the ratio). "
        "NB: 2026-YTD is only ~8 months → annualised figures are noisy. "
        "MC15y med/p5/band = Monte-Carlo 15-year CAGR median, 5th-pct (bad case) and band, "
        "resampled on MONTHLY returns (daily resampling inflated the band with non-synchronous "
        "pricing noise and made it depend on the block length); any MC drawdown is therefore a "
        "MONTH-END one, shallower than and not comparable to the realised daily MaxDD above. "
        "(p95−p5 = OUTCOME DISPERSION; narrower = more reliable, less path-dependent). DD str p5 = "
        "5th-pct 15y drawdown under a GARCH(1,1)+filtered-historical stress, which CAN string "
        "together a persistent high-vol regime the block bootstrap cannot; it moves the drawdown "
        "tail (-33% to -38% on the lead target) and leaves the return distribution alone. corr→1 ΔDD = "
        "extra 5th-pct drawdown if the sleeves stopped diversifying, i.e. how much of the "
        "portfolio rests on correlations that held in the sample but have no law behind them "
        "(-19 points on the lead target); both legs come from the same sleeve-level engine, which "
        "unlike the blended one also actually rebalances. cond CAGR = the SAME sleeve engine "
        "started from today's short rate, long yield and CAPE instead of the sample's average "
        "day, i.e. a scenario under stated assumptions rather than a forecast; the biggest single "
        "driver is assuming gold reverts to ~1% real rather than repeating its realised 10%/yr. "
        "€100k→15y "
        "= €100k compounded 15y at the MC median (your horizon). R5y/R10y/R15y med & p5 = median "
        "and 5th-pct annualised return over all historical rolling 5/10/15y holds (15y windows "
        "few & overlapping → thin, prefer MC15y). Eq/Bond/Gold/Trend/Comm = notional %% of "
        "capital; USA/DevexUS/EMU/JP/EM = %% of the equity sleeve. "
        "ROWS: targets (★) first, then ranked by MC15y median descending, ties broken by "
        "the narrower MC band."
    )).font = _font(8, italic=True, color=_C["muted"])

    gr, hr = 3, 4                                   # group-header row, metric-label row

    def _fixed_hdr(col0, items):
        for k, (lbl, _w) in enumerate(items):
            ws.merge_cells(start_row=gr, start_column=col0 + k, end_row=hr, end_column=col0 + k)
            c = ws.cell(row=gr, column=col0 + k, value=lbl)
            c.font = _font(9, bold=True, color=_C["white"]); c.fill = _fill(_C["header"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
            lo = ws.cell(row=hr, column=col0 + k)
            lo.fill = _fill(_C["header"]); lo.border = _border()

    _fixed_hdr(1, fixed_pre)
    _fixed_hdr(1 + len(fixed_pre), decision)
    wcol = 1 + len(fixed_pre) + len(decision)      # first per-window column
    win_starts = [wcol + gi * len(_WIN_METRICS) for gi in range(len(_SUMMARY_WINDOWS))]
    win_end = wcol + len(_WIN_METRICS) * len(_SUMMARY_WINDOWS) - 1
    for gi, (glabel, _start) in enumerate(_SUMMARY_WINDOWS):
        c0 = win_starts[gi]
        ws.merge_cells(start_row=gr, start_column=c0, end_row=gr,
                       end_column=c0 + len(_WIN_METRICS) - 1)
        gc = ws.cell(row=gr, column=c0, value=glabel)
        gc.font = _font(9, bold=True, color=_C["white"]); gc.fill = _fill(_C["band"])
        gc.alignment = _align("center"); gc.border = _border()
        for mi, (mlbl, _k, _f) in enumerate(_WIN_METRICS):
            mc = ws.cell(row=hr, column=c0 + mi, value=mlbl)
            mc.font = _font(8, bold=True, color=_C["white"]); mc.fill = _fill(_C["header"])
            mc.alignment = _align("center"); mc.border = _border()
    _fixed_hdr(win_end + 1, fixed_post)
    ws.row_dimensions[gr].height = 15
    ws.row_dimensions[hr].height = 20
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)

    CLS = ["Equities", "Fixed Income", "Gold", "Alternative", "Commodities"]
    GEO = ["USA", "Dev ex-USA ex-EMU ex-JP", "Eurozone EMU", "Japan", "Emerging Markets"]

    def _mc_cagr(p) -> dict:
        return ((((p.rob or {}).get("horizons", {}) or {}).get(15, {}) or {})
                .get("mc", {}).get("cagr", {}) or {})

    def _rank_key(p):
        """Targets first, then everything else; within each group MC median
        descending, ties broken by the narrower MC band."""
        mc = _mc_cagr(p)
        med = _clean_num(mc.get("median"))
        p5, p95 = _clean_num(mc.get("p05")), _clean_num(mc.get("p95"))
        band = (p95 - p5) if (p5 is not None and p95 is not None) else None
        # -med sorts descending; +band sorts ascending (tighter dispersion wins a
        # tie). Missing values sink to the bottom rather than poisoning the sort.
        return (0 if p.name in targets else 1,
                -(med if med is not None else -1e9),
                band if band is not None else 1e9)

    ordered = sorted(portfolios, key=_rank_key)
    r = hr + 1
    for i, p in enumerate(ordered):
        is_t = p.name in targets
        # No dark row-fill: the per-column colour scale below owns the data cells,
        # so text stays dark/readable; targets are marked by the ★ + bold name.
        bg = _C["alt"] if i % 2 else _C["white"]
        namecolor = _C["band"] if is_t else _C["text"]
        mccg = _mc_cagr(p)
        mc_med = _clean_num(mccg.get("median"))
        mc_p5 = _clean_num(mccg.get("p05")); mc_p95 = _clean_num(mccg.get("p95"))
        mc_band = (mc_p95 - mc_p5) if (mc_p5 is not None and mc_p95 is not None) else None
        fv = 100_000 * (1.0 + mc_med / 100.0) ** 15 if mc_med is not None else None
        comp = " · ".join(f"{tk} {w:.0f}" for tk, w in
                          sorted(p.weights().items(), key=lambda kv: -kv[1]))
        vals = [("★ " if is_t else "") + p.name.replace("_", " ")]
        # Decision block: MC projection, €100k outcome, rolling holds, allocation.
        vals.append((mc_med, "0.0")); vals.append((mc_p5, "0.0"))
        vals.append((mc_band, "0.0"))
        # GARCH-FHS drawdown tail: the bad case the block bootstrap cannot build.
        vals.append((_clean_num((((p.rob or {}).get("mc_stress", {}) or {})
                                .get("max_drawdown", {}) or {}).get("p05")), "0.0"))
        # Diversification dependence: the extra tail drawdown you take on if the
        # sleeves stop diversifying. Both legs come from the SAME sleeve-level
        # engine, so the difference isolates the correlations.
        _jh = (((p.rob or {}).get("mc_joint", {}) or {}).get("max_drawdown", {}) or {}).get("p05")
        _j1 = (((p.rob or {}).get("mc_joint_corr1", {}) or {}).get("max_drawdown", {}) or {}).get("p05")
        vals.append((_clean_num((_j1 - _jh) if (None not in (_jh, _j1)) else None), "0.0"))
        # Same sleeve engine, started from TODAY's yields and valuations instead
        # of the sample's average day. A scenario under stated assumptions, not a
        # forecast — the gold anchor alone drives over half of it.
        vals.append((_clean_num((((p.rob or {}).get("mc_conditional", {}) or {})
                                .get("cagr", {}) or {}).get("median")), "0.0"))
        vals.append((fv, "eur"))
        vals.append(((fv / 100_000) if fv is not None else None, "0.00\"x\""))  # multiplier
        for hz in (5, 10, 15):                         # rolling historical N-year holds
            roll = (((p.rob or {}).get("horizons", {}) or {}).get(hz, {}) or {}).get("rolling", {})
            for key in ("median", "p05"):
                rv = roll.get(key)
                vals.append((_clean_num(rv * 100.0 if rv is not None else None), "0.0"))
        for key in CLS:
            vals.append((p.notl_gross.get(key, 0.0), "0"))
        for key in GEO:
            vals.append((p.geo_notl.get(key, 0.0), "0"))
        vals.append((p.leverage, "0.00\"x\""))
        for _glabel, start in _SUMMARY_WINDOWS:       # 8 KPIs per window
            wm = _wm(p.synth_nav, start)
            cg = _clean_num(wm.get("cagr")); dd = _clean_num(wm.get("max_drawdown"))
            rowm = {"cagr": cg, "volatility": _clean_num(wm.get("volatility")),
                    "sharpe": _clean_num(wm.get("sharpe")), "sortino": _clean_num(wm.get("sortino")),
                    "max_drawdown": dd, "calmar": (cg / abs(dd)) if (cg is not None and dd) else None,
                    "beta": _clean_num(wm.get("beta")), "alpha": _clean_num(wm.get("alpha"))}
            for _lbl, key, fmt in _WIN_METRICS:
                vals.append((rowm.get(key), fmt))
        vals.append((comp, "s"))
        for j, item in enumerate(vals):
            if j == 0:
                c = ws.cell(row=r, column=1, value=item)
                c.font = _font(9, bold=True, color=namecolor); c.alignment = _align("left")
            elif item[1] == "s":
                c = ws.cell(row=r, column=1 + j, value=item[0])
                c.font = _font(8, color=_C["text"]); c.alignment = _align("left")
            else:
                v, fmt = item
                c = ws.cell(row=r, column=1 + j,
                            value=("n/a" if v is None else eur_smart(v) if fmt == "eur" else v))
                if v is not None and fmt != "eur":
                    c.number_format = fmt
                c.font = _font(9, bold=is_t, color=_C["text"]); c.alignment = _align("center")
            c.fill = _fill(bg); c.border = _border()
        r += 1
    last_data = r - 1

    # Per-column colour scale (best→worst), direction-aware: higher-is-better for
    # CAGR/Sharpe/Sortino/Calmar/α/MaxDD(=less negative)/MC/rolling; lower-is-better
    # for Vol / MC band (dispersion) / β. Allocation, geography, €100k and text are
    # left uncoloured (no intrinsic "better").
    from openpyxl.formatting.rule import ColorScaleRule
    _GRN, _YEL, _RED = "63BE7B", "FFEB84", "F8696B"

    def _cs(higher_better):
        a, z = (_RED, _GRN) if higher_better else (_GRN, _RED)
        return ColorScaleRule(start_type="min", start_color=a, mid_type="percentile",
                              mid_value=50, mid_color=_YEL, end_type="max", end_color=z)

    def _blue():   # magnitude heatmap: low = white, high = dark blue
        return ColorScaleRule(start_type="min", start_color="FFFFFF",
                              end_type="max", end_color="1D4ED8")

    per_win = ["hi", "lo", "hi", "hi", "hi", "hi", "lo", "hi"]  # CAGR Vol Shrp Sort MaxDD Calmr β α
    base = 1 + len(fixed_pre)                              # MC15y med column
    # MC med, p5, band, then the GARCH drawdown tail (less negative = better).
    dirs = {base: "hi", base + 1: "hi", base + 2: "lo", base + 3: "hi",
            base + 4: "hi", base + 5: "hi"}
    dirs[base + 7] = "hi"                                  # ×mult (base+6 = €100k text, skip)
    for k in range(8, 14):                                 # R5/R10/R15 med & p5
        dirs[base + k] = "hi"
    for wi in range(len(_SUMMARY_WINDOWS)):                # the per-window KPI blocks
        for mi, d in enumerate(per_win):
            dirs[win_starts[wi] + mi] = d
    for col, d in dirs.items():
        L = get_column_letter(col)
        ws.conditional_formatting.add(f"{L}{hr + 1}:{L}{last_data}", _cs(d == "hi"))
    # Magnitude heatmap (white→dark blue) on asset split + geography + leverage.
    lev_col = base + len(decision) - 1
    for col in list(range(base + 14, base + 24)) + [lev_col]:   # Eq..EM + Lev
        L = get_column_letter(col)
        ws.conditional_formatting.add(f"{L}{hr + 1}:{L}{last_data}", _blue())

    # Vertical rules boxing each time window, so the four KPI blocks read as four
    # tables instead of 32 undifferentiated columns. Applied over the header rows
    # and every data row, preserving the existing thin grid on the other sides.
    def _vrule(col, side):
        thick = Side(style="medium", color=_C["header"])
        thin = Side(style="thin", color=_C["border"])
        for rr in range(gr, last_data + 1):
            c = ws.cell(row=rr, column=col)
            c.border = Border(left=thick if side == "left" else thin,
                              right=thick if side == "right" else thin,
                              top=thin, bottom=thin)

    for c0 in win_starts:                                  # left edge of each window
        _vrule(c0, "left")
    # Close the last window on the LEFT edge of the next column rather than the
    # right edge of its own: that column's header is a vertical merge whose style
    # lives on the anchor cell, so the rule renders on the header row too.
    _vrule(win_end + 1, "left")

    # Charts, stacked below the table: (1) 4-window growth (log), all portfolios;
    # (2) Monte-Carlo 15y fan for each target.
    try:
        from openpyxl.drawing.image import Image as XLImage

        def _place(buf, at_row):
            """Anchor an image and return the first free row BELOW it.

            The row span is derived from the image's own pixel height (a default
            Excel row is 20px) instead of a hand-tuned constant, so a chart that
            grows — a taller legend, another panel — can never overlap the next
            one.
            """
            img = XLImage(buf)
            ws.add_image(img, f"A{at_row}")
            return at_row + int(img.height / 20) + 2

        r += 1
        gh = ws.cell(row=r, column=1, value="Cumulative return — 4 windows (log), all portfolios")
        gh.font = _font(11, bold=True, color=_C["white"]); gh.fill = _fill(_C["band"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(ncol, 12))
        r += 1
        gbuf = _growth_windows_png(portfolios, metrics_of=_wm)
        if gbuf is not None:
            r = _place(gbuf, r)
        r += 2
        h = ws.cell(row=r, column=1,
                    value="Monte-Carlo 15-year projection (block-bootstrap of the history, €100k)")
        h.font = _font(11, bold=True, color=_C["white"]); h.fill = _fill(_C["band"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(ncol, 12))
        r += 1
        for p in ordered:                              # same order as the table
            if p.name not in targets:
                continue
            buf = _mc_fan_png(getattr(p, "synth_nav", None), p.name.replace("_", " "))
            if buf is not None:
                r = _place(buf, r)
    except Exception:  # noqa: BLE001
        pass


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_whatif_excel(path, portfolios, asset_target, geo_target, anchor,
                        tolerance=1.5, sim_rows=None, testfol=None,
                        testfol_byinst=None) -> str:
    """Write the single-sheet, column-per-portfolio comparison workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "What-If"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["band"]

    ncol = _PCOL0 + len(portfolios)  # incl. a possible Target column
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 44
    # Per-portfolio columns sized to fit a "100.0%" cell and let the wrapped
    # header name break at word boundaries (fit-to-content, not oversized).
    for j in range(len(portfolios) + 1):
        ws.column_dimensions[get_column_letter(_PCOL0 + j)].width = 10

    # Title.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value="WHAT-IF PORTFOLIO COMPARISON")
    t.font = _font(16, bold=True, color=_C["white"])
    t.fill = _fill(_C["header"])
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    st = ws.cell(row=2, column=1, value=(
        f"Notional anchor (display only): {eur_smart(anchor)}  ·  "
        f"generated {datetime.now():%Y-%m-%d %H:%M}"))
    st.font = _font(9, italic=True, color=_C["muted"])

    # Four sheets, no more: Summary (one row per portfolio, the decision table),
    # What-If (this one: instruments × portfolios, diversification by asset class
    # and geography, gross/leverage, risk metrics EUR/USD), Simulation (how each
    # instrument's pre-inception history was backfilled) and Testfolio. The former
    # Robustness and Horizons sheets are gone: everything an allocation decision
    # needs from them (MC 15y median/p5/band, rolling 5/10/15y) is a column in
    # Summary, so a whole sheet per view was duplication.
    row = 4
    row = _lineage_block(ws, row, portfolios, ncol)
    row = _instrument_matrix_block(ws, row, portfolios)

    # --- Diversification: asset class NOTIONAL exposure (gross, leverage-aware),
    # coloured by drift vs target — like the newsletter's "By asset class".
    gross_labels = [c for c in _ASSET_ORDER if any(p.notl_gross.get(c, 0.0) for p in portfolios)]
    gross_tbl = _alloc_matrix(
        ws, row, "Diversification — by asset class (NOTIONAL, % of capital) vs target",
        gross_labels, portfolios, lambda p, l: p.notl_gross.get(l, 0.0),
        target=asset_target, tolerance=tolerance, swatch=asset_class_color)
    row = gross_tbl["last"] + 1
    tgt_gross = sum(v for v in (asset_target or {}).values() if v) or None
    row = _plain_row(ws, row, "Gross exposure", portfolios,
                     lambda p: f"{p.gross:.0f}%")
    row = _plain_row(ws, row, "Leverage", portfolios,
                     lambda p: f"{p.leverage:.2f}x", color=_C["amber"])
    row += 1

    # Leverage by class = notional / funded capital in each class (>1.00x =
    # partly synthetic). Shown as a ratio (like the newsletter), only when some
    # class is levered.
    def _lev(p, c):
        # cap == 0 with notional > 0 = fully synthetic exposure (e.g. FI held
        # only via NTSG's futures overlay) → ∞, NOT "absent".
        notl = p.notl_gross.get(c, 0.0)
        cap = p.cap.get(c, 0.0)
        if cap > 0:
            return notl / cap
        return float("inf") if notl > 0 else None
    lev_labels = [c for c in _ASSET_ORDER
                  if any((_lev(p, c) or 0) > 1.001 for p in portfolios)]
    if lev_labels:
        ncols = _PCOL0 + len(portfolios) - 1
        row = _section_header(ws, row, "Diversification — leverage by class (notional / funded)", ncols)
        _col_headers(ws, row, portfolios, "Category", with_target=False)
        row += 1
        for k, cls in enumerate(lev_labels):
            bg = _C["alt"] if k % 2 else _C["white"]
            _merge_label(ws, row, cls, bold=True, color=asset_class_color(cls), bg=bg)
            for j, p in enumerate(portfolios):
                lv = _lev(p, cls)
                c = ws.cell(row=row, column=_PCOL0 + j,
                            value=("\u2014" if lv is None
                                   else "\u221e (all synthetic)" if lv == float("inf")
                                   else f"{lv:.2f}x"))
                c.alignment = _align("center")
                c.fill = _fill(bg)
                c.border = _border()
                c.font = _font(9)
            row += 1
        row += 1

    geo_labels = [c for c in _GEO_ORDER if any(p.geo_notl.get(c, 0.0) for p in portfolios)]
    geo_tbl = _alloc_matrix(
        ws, row, "Diversification — by geography (NOTIONAL, % of equity sleeve) vs target",
        geo_labels, portfolios, lambda p, l: p.geo_notl.get(l, 0.0),
        target=geo_target, tolerance=tolerance, swatch=geo_color)
    row = geo_tbl["last"] + 2

    row = _risk_matrix(ws, row, portfolios)

    if sim_rows:
        _simulation_sheet(wb, sim_rows)
    if testfol:
        _testfol_sheet(wb, testfol, testfol_byinst)
    _method_sheet(wb, portfolios)

    # Summary FIRST (created at index 0): one row per portfolio, full metric set.
    _summary_sheet(wb, portfolios)

    wb.save(path)
    logger.info("What-if workbook saved to %s", path)
    return path


# ---------------------------------------------------------------------------
# Method & assumptions
# ---------------------------------------------------------------------------

# Every column in the Summary, and what it rests on. Kept as data rather than
# prose in a docstring so the workbook can explain itself to a reader who does
# not have the code — the numbers are only usable if the assumptions travel with
# them. (group, column, what it is, how it is computed, what it assumes).
_METHOD_ROWS: list[tuple[str, str, str, str]] = [
    ("ENGINE", "Synthetic history",
     "One aligned daily series per portfolio, 2000-2026, in EUR.",
     "Each sleeve is rebuilt from geo/market proxies (leverage financed at the short rate + 0.5% spread), "
     "net of TER, then spliced onto the fund's real returns once they start. Factor ETFs get a factor tilt "
     "on the Fama-French legs. ASSUMES the proxies represent the sleeve and that a reconstructed past is "
     "informative; the pre-inception tail omits the regression residual, so its idiosyncratic vol is a floor."),
    ("ENGINE", "Rebalancing",
     "Quarterly by default: weights drift between dates and reset at each boundary.",
     "Modelled with no transaction cost — no commissions, no bid-ask, no tax. On 8 positions at ~EUR19/order "
     "that is 0.13-0.26%/yr of real drag not shown anywhere in this workbook. The sleeve simulation below "
     "measures rebalancing as worth ~0.19pp of CAGR against 1.1pp of dispersion, with annual ~= quarterly."),
    ("ENGINE", "Currency & inflation",
     "Everything is NOMINAL EUR.",
     "No inflation adjustment anywhere. At 2% inflation a 15-year nominal multiplier of 3.4x is 2.5x in "
     "today's purchasing power. Sharpe/Sortino/Calmar/MaxDD/beta are unaffected (nominal excess over a "
     "nominal risk-free is already a real quantity); the wealth projections are not."),
    ("RETURN", "CAGR",
     "Annualised compound return over the window, from endpoint to endpoint.",
     "Frequency-independent, so unaffected by the daily-noise issue below. Net of TER and financing, gross "
     "of commissions and tax."),
    ("RISK", "Vol%",
     "Annualised standard deviation, from MONTHLY returns.",
     "Monthly on purpose: the reconstruction's DAILY frequency carries non-synchronous pricing noise "
     "(proxies closing in different time zones, an FX stamp at another hour, stale historical NAVs) which "
     "inflated vol by ~1.4x. Proof it is noise: over 2021+, where the funds have real prices, daily and "
     "monthly agree (ratio 1.06 vs 1.41 full-window). It hit diversified portfolios hardest (1.43x on 8 "
     "sleeves vs 1.24x on 2), because noise is independent across sleeves while true returns are not."),
    ("RISK", "Shrp / Sort",
     "Sharpe and Sortino on monthly excess returns over a time-varying risk-free.",
     "Each month is charged the rate that actually prevailed. Sortino uses target semideviation (RMS "
     "shortfall below the risk-free), not the std of negatives. Both were ~30-40% understated while "
     "measured daily."),
    ("RISK", "MaxDD% / Calmr",
     "Worst peak-to-trough on the DAILY path; Calmar = CAGR / |MaxDD|.",
     "Deliberately daily, unlike Vol: a month-end drawdown would erase real intra-month crashes such as "
     "March 2020. Not comparable to the Monte-Carlo drawdowns, which are month-end."),
    ("RISK", "beta / alpha",
     "Daily regression against an ACWI-weighted blend of the geo proxies.",
     "Kept daily because portfolio and benchmark are built from the same proxies, so their timing noise is "
     "common and cancels in the ratio (beta moves 0.71 -> 0.70 monthly) while daily gives 20x the "
     "observations. Alpha is Jensen, annualised."),
    ("MONTE CARLO", "MC15y med / p5 / band",
     "Median, 5th percentile and p95-p5 spread of the 15-year CAGR over 2000 simulated paths.",
     "Moving-block bootstrap of MONTHLY returns, 1-month blocks: the portfolio's own months are drawn with "
     "replacement and compounded into new 15-year paths. It reshuffles the PAST, it does not forecast. "
     "Monthly and 1-month blocks because the answer then stops depending on the block length (band 10.1 / "
     "10.1 / 10.1 / 10.5 for 1/3/6/12-month blocks, against 14.4 / 11.7 / 10.6 for daily 1/21/126-day "
     "blocks). ASSUMES the future is drawn from the sample, with its means, correlations and tails."),
    ("MONTE CARLO", "DD str p5",
     "5th-percentile 15-year drawdown under a volatility-clustering stress.",
     "GARCH(1,1) fitted to monthly returns, then the STANDARDISED residuals are resampled (filtered "
     "historical simulation), so shock shapes stay empirical. Exists because the bootstrap draws months "
     "independently and structurally cannot build a persistent bad regime, which is what deep drawdowns are "
     "made of. It moves the tail only (-33% -> -38% on the lead target) and leaves returns alone, so quote "
     "the bootstrap's CAGR, not this one's."),
    ("MONTE CARLO", "corr->1 dDD",
     "Extra 5th-percentile drawdown if the sleeves stopped diversifying.",
     "Sleeve-level simulation twice: once at historical correlations, once with a Gaussian copula that keeps "
     "each sleeve's empirical monthly marginal while sliding the correlation matrix to all-ones. The "
     "difference is how much of the portfolio rests on correlations that held in the sample but have no law "
     "behind them. Both legs share one engine, so nothing else differs."),
    ("MONTE CARLO", "cond CAGR",
     "15-year median CAGR started from TODAY's yields and valuations instead of the sample's average day.",
     "A SCENARIO UNDER STATED ASSUMPTIONS, NOT A FORECAST. Only the means move; volatilities, correlations "
     "and tails stay empirical. Shifts scale by each fund's own exposure, so a 90/60 efficient-core takes "
     "0.9 of the equity shift and 0.6 of the bond shift and a 2x ETF takes double. See the driver table "
     "below for what moves each portfolio. TWO LIMITS: the equity anchor is the US CAPE applied to ALL "
     "equity sleeves, so it overstates the drag on the cheaper non-US ones; and because the only views are "
     "per asset class, portfolios with the same class mix get the SAME shift — this column separates a "
     "gold-heavy portfolio from a bond-heavy one, but says nothing about which equities you hold."),
    ("HISTORY", "R5y / R10y / R15y",
     "Median and 5th percentile annualised return over every rolling 5/10/15-year hold in the history.",
     "Overlapping windows, so they are not independent observations: 26 years hold about 2-3 independent "
     "15-year periods. Read them as texture, prefer the Monte Carlo for the horizon."),
    ("EXPOSURE", "Lev / Eq..Comm / USA..EM",
     "Leverage, notional asset split as % of capital, equity geography as % of the equity sleeve.",
     "Notional, i.e. leverage-aware: an efficient-core fund contributes both its equity and its bond "
     "futures. Geography comes from the taxonomy's per-instrument breakdown, not from a lookthrough."),
    ("PROJECTION", "EUR100k->15y / xmult",
     "The MC median CAGR compounded for 15 years from EUR100k.",
     "Median, not mean: compound returns are right-skewed, so the mean is dragged up by a few paths nobody "
     "will live (EUR389k mean vs EUR352k median on the lead target). NOMINAL, and it inherits every "
     "assumption of the MC column it comes from."),
]


def _method_sheet(wb, portfolios) -> None:
    """A sheet that explains the workbook: what each Summary column is, how it is
    computed and what it assumes, followed by an additive attribution of the
    conditional scenario per portfolio.

    A number without its assumptions is not usable. The conditional CAGR in
    particular is a scenario, and a reader who cannot see that gold contributes
    more of it than the financing cost does has no way to argue with it — which
    is the only thing that earns a scenario a place in a decision.
    """
    ws = wb.create_sheet("Method")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["muted"]
    for col, wid in (("A", 14), ("B", 24), ("C", 58), ("D", 112)):
        ws.column_dimensions[col].width = wid

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="METHOD & ASSUMPTIONS — how every Summary column is built")
    t.font = _font(15, bold=True, color=_C["white"]); t.fill = _fill(_C["header"])
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value=(
        "Read this before quoting anything from Summary. Backtested figures are net of TER and financing "
        "and GROSS of commissions and tax; everything is nominal EUR. Where a choice was made — sampling "
        "frequency, an anchor, a haircut — it is named here rather than buried."
    )).font = _font(9, italic=True, color=_C["muted"])
    ws.row_dimensions[2].height = 24

    hdr = 3
    for j, lab in enumerate(("Block", "Column", "What it is",
                             "How it is computed / what it assumes")):
        c = ws.cell(row=hdr, column=1 + j, value=lab)
        c.font = _font(9, bold=True, color=_C["white"]); c.fill = _fill(_C["band"])
        c.alignment = _align("left"); c.border = _border()
    r = hdr + 1
    last_group = None
    for group, column, what, how in _METHOD_ROWS:
        bg = _C["alt"] if (r % 2) else _C["white"]
        for j, val in enumerate((group if group != last_group else "", column, what, how)):
            c = ws.cell(row=r, column=1 + j, value=val)
            c.font = _font(9, bold=(j == 0 and bool(val)), color=_C["text"])
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.fill = _fill(bg); c.border = _border()
        ws.row_dimensions[r].height = max(28, 11 * (1 + len(how) // 100))
        last_group = group
        r += 1

    def _first(key):
        for q in portfolios:
            v = ((q.rob or {}).get("mc_conditional", {}) or {}).get(key)
            if v:
                return v
        return None

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    h = ws.cell(row=r, column=1,
                value="CONDITIONAL SCENARIO — what moves each portfolio (additive attribution)")
    h.font = _font(12, bold=True, color=_C["white"]); h.fill = _fill(_C["header"])
    r += 1
    state, comp = _first("state"), _first("components")
    bits = []
    if state:
        for lab, kn, ka, fmt in (("short rate", "short_rate_now", "short_rate_avg", "{:.2f}%"),
                                 ("10y yield", "long_yield_now", "long_yield_avg", "{:.2f}%"),
                                 ("US CAPE", "cape_now", "cape_avg", "{:.1f}")):
            if state.get(kn) is not None and state.get(ka) is not None:
                bits.append(f"{lab} today {fmt.format(state[kn])} vs sample avg {fmt.format(state[ka])}")
        if state.get("cape_asof"):
            bits.append(f"CAPE as of {state['cape_asof']}")
    if comp:
        if comp.get("equity") is not None:
            bits.append(
                f"equity shift {comp['equity']:+.2f}pp/yr (forward "
                f"{comp.get('equity_forward') or 0:+.2f} minus in-sample "
                f"{comp.get('equity_in_sample') or 0:+.2f}, so the sample's own valuation move is "
                f"credited back instead of charged twice)")
        if comp.get("fixed_income") is not None:
            bits.append(f"bond shift {comp['fixed_income']:+.2f}pp/yr")
        bits.append(f"gold assumed {comp.get('gold_target') or 0:.1f}% nominal vs what it realised")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value=("STARTING POINT — " + "  |  ".join(bits) if bits
                                    else "Conditional scenario unavailable.")
            ).font = _font(9, italic=True, color=_C["muted"])
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value=(
        "Each driver = that class's shift times this portfolio's exposure to the class, so the four add up. "
        "The %-of-total columns show WHICH assumption is doing the work: where gold dominates, the number is "
        "mostly a judgement about gold rather than about the portfolio. 'total pp' is the analytic sum; the "
        "simulated gap in Summary differs from it by rebalancing and compounding."
    )).font = _font(9, italic=True, color=_C["muted"])
    ws.row_dimensions[r].height = 30
    r += 2

    # --- How each shift is computed, with the live inputs -------------------
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    hh = ws.cell(row=r, column=1, value="SHIFT ARITHMETIC — every number below is reproducible from these")
    hh.font = _font(11, bold=True, color=_C["white"]); hh.fill = _fill(_C["band"])
    r += 1
    for j, lab in enumerate(("Driver", "Formula", "Inputs (live)", "Result")):
        c = ws.cell(row=r, column=1 + j, value=lab)
        c.font = _font(9, bold=True, color=_C["white"]); c.fill = _fill(_C["header"])
        c.alignment = _align("left"); c.border = _border()
    r += 1
    sy = _first("sample_years") or 25.0
    gold_t = (comp or {}).get("gold_target")
    arith_rows = [
        ("equity valuation",
         "shift = forward - in_sample, where each = (CAPE_to / CAPE_from) ^ (1 / years) - 1",
         (f"in_sample: CAPE {state.get('cape_at_sample_start'):.1f} -> {state.get('cape_now'):.1f} "
          f"over {sy:.1f}y = {(comp or {}).get('equity_in_sample') or 0:+.2f}%/yr  |  "
          f"forward: CAPE {state.get('cape_now'):.1f} -> {state.get('cape_avg'):.1f} over 15y = "
          f"{(comp or {}).get('equity_forward') or 0:+.2f}%/yr"
          if state and comp else "n/a"),
         f"{(comp or {}).get('equity') or 0:+.2f} pp/yr"),
        ("", "WHY the subtraction",
         "The sample's realised return already contains whatever the multiple did INSIDE it. Charging the "
         "forward reversion alone would bill that move twice. Here the sample started and ended at almost "
         "the same multiple, so it carries almost no drag to credit back and the forward figure passes "
         "through nearly whole.", ""),
        ("bond yield",
         "shift = yield_today - average yield over the sample",
         (f"10y {state.get('long_yield_now'):.2f}% today vs {state.get('long_yield_avg'):.2f}% average"
          if state and state.get("long_yield_now") is not None else "n/a"),
         f"{(comp or {}).get('fixed_income') or 0:+.2f} pp/yr"),
        ("", "WHY the starting yield",
         "For a bond held near its duration the entry yield is the dominant term of the eventual return, "
         "so today's level is a genuine anchor rather than an assumption. It is the one driver here that "
         "helps: yields are higher now than they averaged.", ""),
        ("gold",
         "shift = (assumed real return + assumed inflation) - what this sleeve REALISED in the sample",
         (f"assumed {gold_t:.1f}% nominal (1% real + 2% inflation) against a realised rate that ran above "
          f"10%/yr in this window" if gold_t is not None else "n/a"),
         "per sleeve (see table)"),
        ("", "WHY an absolute level",
         "Gold pays no cash flow, so no yield or multiple anchors it and its realised rate is not an "
         "expectation. This is a JUDGEMENT, and it is the assumption to argue with first: moving the real "
         "return from 1% to 4% lifts the lead target's conditional CAGR by about half a point.", ""),
        ("financing",
         "shift = -(short rate today - average short rate) x borrowed notional",
         (f"{state.get('short_rate_now'):.2f}% today vs {state.get('short_rate_avg'):.2f}% average, "
          f"borrowed = leverage - 1"
          if state and state.get("short_rate_now") is not None else "n/a"),
         "per portfolio (see table)"),
        ("", "WHY on every sleeve",
         "The backtest financed leverage at the rate that prevailed then. Higher now means every sleeve's "
         "return has more to clear, including the ones with no valuation view — otherwise trend and "
         "commodities would ride the leverage for free.", ""),
        ("per sleeve",
         "shift_sleeve = SUM over classes ( exposure_to_class x shift_class ) + financing",
         "Exposure is per unit of the FUND's own NAV and is not normalised: a 90/60 efficient-core takes "
         "0.9 x equity + 0.6 x bond, a 2x ETF takes 2 x equity.", "input to the simulation"),
        ("per driver",
         "contribution = shift_class x SUM over sleeves ( portfolio weight x exposure_to_class )",
         "This is why the four columns add up, and why two portfolios with the same class mix show the same "
         "attribution even when they hold different funds.", "the table below"),
    ]
    for i, (drv, formula, inputs, res) in enumerate(arith_rows):
        bg = _C["alt"] if i % 2 else _C["white"]
        for j, val in enumerate((drv, formula, inputs, res)):
            c = ws.cell(row=r, column=1 + j, value=val)
            c.font = _font(9, bold=(j == 0 and bool(drv)),
                           italic=formula.startswith("WHY"), color=_C["text"])
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.fill = _fill(bg); c.border = _border()
        ws.row_dimensions[r].height = max(26, 11 * (1 + max(len(inputs), len(formula)) // 95))
        r += 1
    r += 2

    cols = [("Portfolio", 24), ("MC15y med", 10), ("cond CAGR", 10), ("total pp", 9),
            ("equity pp", 10), ("bond pp", 9), ("gold pp", 9), ("fin pp", 8),
            ("equity % of total", 15), ("gold % of total", 14)]
    for j, (lab, wid) in enumerate(cols):
        c = ws.cell(row=r, column=1 + j, value=lab)
        c.font = _font(9, bold=True, color=_C["white"]); c.fill = _fill(_C["band"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        if j:
            ws.column_dimensions[get_column_letter(1 + j)].width = wid
    ws.row_dimensions[r].height = 26
    r += 1
    first_data = r
    ordered = sorted(
        portfolios,
        key=lambda q: -((((q.rob or {}).get("mc_conditional", {}) or {})
                         .get("cagr", {}) or {}).get("median") or -1e9))
    for i, p in enumerate(ordered):
        cond = ((p.rob or {}).get("mc_conditional", {}) or {})
        dr = cond.get("drivers") or {}
        if not dr:
            continue
        mc = ((((p.rob or {}).get("horizons", {}) or {}).get(15, {}) or {})
              .get("mc", {}).get("cagr", {}) or {}).get("median")
        gross = sum(abs(v) for v in dr.values()) or 1.0
        bg = _C["alt"] if i % 2 else _C["white"]
        vals = [("★ " if p.name in _TARGETS else "") + p.name.replace("_", " "),
                (_clean_num(mc), "0.00"),
                (_clean_num((cond.get("cagr", {}) or {}).get("median")), "0.00"),
                (_clean_num(cond.get("drivers_total")), "0.00"),
                (_clean_num(dr.get("equity_valuation")), "0.00"),
                (_clean_num(dr.get("bond_yield")), "0.00"),
                (_clean_num(dr.get("gold")), "0.00"),
                (_clean_num(dr.get("financing")), "0.00"),
                (abs(dr.get("equity_valuation") or 0.0) / gross, "0%"),
                (abs(dr.get("gold") or 0.0) / gross, "0%")]
        for j, item in enumerate(vals):
            if j == 0:
                c = ws.cell(row=r, column=1, value=item)
                c.font = _font(9, bold=p.name in _TARGETS, color=_C["text"])
                c.alignment = _align("left")
            else:
                v, fmt = item
                c = ws.cell(row=r, column=1 + j, value=("n/a" if v is None else v))
                if v is not None:
                    c.number_format = fmt
                c.font = _font(9, color=_C["text"]); c.alignment = _align("center")
            c.fill = _fill(bg); c.border = _border()
        r += 1
    # --- The four questions a reader actually asks ---------------------------
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    hq = ws.cell(row=r, column=1, value="READING IT CORRECTLY — what the shift is against, and what it is not")
    hq.font = _font(11, bold=True, color=_C["white"]); hq.fill = _fill(_C["band"])
    r += 1
    asof = ""
    if state:
        parts = [f"{lab} {state.get(k)}" for lab, k in
                 (("short rate", "short_rate_asof"), ("10y yield", "long_yield_asof"),
                  ("CAPE", "cape_asof")) if state.get(k)]
        asof = "; ".join(parts)
    qa = [
        ("the shift is against WHAT?",
         "Against each sleeve's OWN REALISED RETURN in the sample, not against the median CAGR. The annual "
         "shift is converted to a monthly one ((1+pp/100)^(1/12)-1) and ADDED to every resampled month of "
         "that sleeve, so it moves the mean of the distribution the bootstrap is drawing from. The median "
         "CAGR is the OUTPUT, and it does not move by the same amount: compounding a lower mean for 180 "
         "months and rebalancing on the way turns an analytic -3.13 pp into a simulated -3.65."),
        ("why is GOLD computed differently?",
         "Equities and bonds have an OBSERVABLE anchor — a CAPE says whether equities are expensive, a yield "
         "says what a bond will pay — so their shift is DERIVED. Gold pays no dividend, coupon or earnings, "
         "so there is no price-to-anything and nothing to derive from: you cannot say gold is expensive the "
         "way you can for an index. So the row states an ABSOLUTE expectation (1% real, from the ~120-year "
         "long-run record where gold roughly preserves purchasing power, plus 2% inflation = 3% nominal) and "
         "subtracts what the sleeve actually delivered. 'Per sleeve' in the Result column means exactly that: "
         "the answer depends on each sleeve's own realised rate, unlike equities and bonds where one class "
         "shift serves all. This is the one row that is entirely a JUDGEMENT — if gold repeated its realised "
         "rate the shift would be ~0 and the lead target's conditional CAGR would be about a point higher."),
        ("what does TODAY mean?",
         "Not one date. The rates are the last TRADING day, the CAPE is the last PUBLISHED MONTH (Shiller is "
         "monthly and lags). As of this workbook: " + (asof or "n/a") + ". 'Valuations' means precisely the "
         "Shiller CAPE on the S&P 500 — real price over the ten-year average of real earnings — which is the "
         "only valuation anchor available here."),
        ("does the ORDER of the months matter?",
         "For the final value, NO: compounding is a product and multiplication commutes, so -50% then +50% "
         "and +50% then -50% both land on 75 from 100. Permuting all the sampled months therefore leaves the "
         "CAGR untouched — measured on this history, a band of exactly 0.00. For the DRAWDOWN, very much yes: "
         "the same three months (-30%, -30%, +60%) give a -51% drawdown when the two losses land back to back "
         "and -30% when a rally separates them, and permuting the real history moved the max drawdown across "
         "a 23-point band at zero CAGR dispersion. That is sequence risk in pure form, and it is why the "
         "drawdown columns have to be simulated rather than derived from the return."),
        ("what about the volatility drag?",
         "It is inherited, not modelled. The simulation never applies an average return: it compounds the "
         "monthly returns one by one, (1+r1)x(1+r2)x...x(1+r180), so the gap between the arithmetic and the "
         "geometric mean is there by construction. On the gold sleeve the arithmetic mean annualises to "
         "+11.69% while the compound rate is +10.29% — a 1.39pp drag, almost exactly the sigma^2/2 of 1.29pp. "
         "Using the arithmetic mean instead would overstate 15-year wealth by roughly a quarter. It also "
         "explains why a shift does not move the answer by its own size: subtracting a constant from every "
         "month leaves volatility untouched while lowering the mean, so the drag weighs relatively more on "
         "what is left. A -7.78pp shift on gold moved its simulated median by -8.52pp, and the same effect is "
         "part of why the analytic attribution below (-3.13) and the simulated gap (-3.65) differ."),
        ("do the MC paths start from today?",
         "NO, and this is the sharpest limitation. The unconditional MC starts from today's MONEY but from "
         "the sample's average CONDITIONS: it draws random months from 2000-2026 and knows nothing about "
         "where we are. The conditional column draws the same months with a shifted mean, which is a CONSTANT "
         "drift whose 15-year average equals what the reversion would deliver — it is NOT a path where the "
         "CAPE actually falls from today's level to the anchor. Three consequences: the reversion is spread "
         "evenly rather than front- or back-loaded, so sequence risk around WHEN multiples compress is not "
         "modelled; there is no feedback, so a crash does not make the next month cheaper and therefore more "
         "promising; and only the mean was moved, so the band was not widened for uncertainty about the "
         "assumption itself. The one piece that genuinely starts from today is the GARCH stress, which begins "
         "at the CURRENT conditional volatility rather than the long-run one. A model that starts from today "
         "in the full sense — state variables evolving and returns following — is a VAR of the Vanguard-VCMM "
         "kind, and is not built here."),
    ]
    for i, (q, a) in enumerate(qa):
        bg = _C["alt"] if i % 2 else _C["white"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=q)
        c.font = _font(9, bold=True, color=_C["band"])
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = _fill(bg); c.border = _border()
        ws.cell(row=r, column=2).fill = _fill(bg)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
        c2 = ws.cell(row=r, column=3, value=a)
        c2.font = _font(9, color=_C["text"])
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.fill = _fill(bg); c2.border = _border()
        for cc in range(4, 11):
            ws.cell(row=r, column=cc).fill = _fill(bg)
        ws.row_dimensions[r].height = max(30, 11 * (1 + len(a) // 150))
        r += 1
    r += 2

    # --- Worked example: the same arithmetic, sleeve by sleeve --------------
    by_name = {q.name: q for q in portfolios}
    lead = next((by_name[n] for n in _TARGETS
                 if n in by_name
                 and ((by_name[n].rob or {}).get("mc_conditional", {}) or {})
                 .get("sleeve_exposures")), None)
    if lead is not None:
        cond = (lead.rob or {}).get("mc_conditional", {}) or {}
        expo = cond.get("sleeve_exposures") or {}
        wts = cond.get("weights") or {}
        shifts = cond.get("drift_shift") or {}
        realised = cond.get("realised") or {}
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        hw = ws.cell(row=r, column=1, value=(
            f"WORKED EXAMPLE — {lead.name.replace('_', ' ')}, sleeve by sleeve "
            f"(borrowed notional {cond.get('leverage_borrowed') or 0:.0%})"))
        hw.font = _font(11, bold=True, color=_C["white"]); hw.fill = _fill(_C["band"])
        r += 1
        heads = ("Sleeve", "Weight", "Equity expo", "Bond expo", "Gold expo",
                 "Realised %/yr", "Shift pp/yr", "x weight = pp", "")
        for j, lab in enumerate(heads):
            c = ws.cell(row=r, column=1 + j, value=lab)
            c.font = _font(9, bold=True, color=_C["white"]); c.fill = _fill(_C["header"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
        r += 1
        tot = 0.0
        for i, tk in enumerate(sorted(expo, key=lambda t: -wts.get(t, 0.0))):
            e = expo[tk]
            wt = wts.get(tk, 0.0)
            sh = shifts.get(tk)
            contrib = (sh * wt) if sh is not None else None
            if contrib is not None:
                tot += contrib
            bg = _C["alt"] if i % 2 else _C["white"]
            row_vals = [tk, (wt, "0%"), (e.get("Equities"), "0.00"),
                        (e.get("Fixed Income"), "0.00"), (e.get("Gold"), "0.00"),
                        (_clean_num(realised.get(tk)), "0.00"),
                        (_clean_num(sh), "0.00"), (_clean_num(contrib), "0.00"), ("", "s")]
            for j, item in enumerate(row_vals):
                if j == 0:
                    c = ws.cell(row=r, column=1, value=item)
                    c.font = _font(9, bold=True, color=_C["text"]); c.alignment = _align("left")
                else:
                    v, fmt = item
                    c = ws.cell(row=r, column=1 + j,
                                value=("" if v in (None, "") else v))
                    if isinstance(v, float):
                        c.number_format = fmt
                    c.font = _font(9, color=_C["text"]); c.alignment = _align("center")
                c.fill = _fill(bg); c.border = _border()
            r += 1
        c = ws.cell(row=r, column=1, value="weighted total")
        c.font = _font(9, bold=True, color=_C["band"]); c.alignment = _align("left")
        c2 = ws.cell(row=r, column=8, value=_clean_num(tot))
        c2.number_format = "0.00"; c2.font = _font(9, bold=True, color=_C["band"])
        c2.alignment = _align("center")
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1, value=(
            "A blank exposure means the sleeve has no view for that class. A sleeve with no valuation "
            "anchor at all (trend, commodities) still carries the financing shift and nothing else. The "
            "weighted total is the analytic answer; the simulation's own median differs from it by "
            "rebalancing and by compounding a shifted mean over 180 months."
        )).font = _font(9, italic=True, color=_C["muted"])
        ws.row_dimensions[r].height = 28
        r += 1

    if r > first_data:
        from openpyxl.formatting.rule import ColorScaleRule
        for col in range(4, 11):
            L = get_column_letter(col)
            ws.conditional_formatting.add(
                f"{L}{first_data}:{L}{r - 1}",
                ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B"))
    # No freeze_panes here on purpose: the driver table starts ~25 rows down, and
    # freezing at it pins more than a screenful, which reads as a sheet that
    # cannot scroll. This is a document, not a grid to navigate.


def _simulation_sheet(wb, sim_rows) -> None:
    """Sheet listing, per instrument, the proxy 'simulation' used to backfill
    its pre-inception history (real fund returns take over from 'Real from')."""
    ws = wb.create_sheet("Simulation")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["green"]
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="SIMULATION MAP — per-instrument history reconstruction")
    t.font = _font(14, bold=True, color=_C["white"])
    t.fill = _fill(_C["header"])
    ws.row_dimensions[1].height = 22

    headers = ["Ticker", "Real from", "Base from", "Simulation base (proxy × exposure)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = _font(9, bold=True, color=_C["white"])
        c.fill = _fill(_C["header"])
        c.border = _border()
    row = 4
    for i, r in enumerate(sim_rows):
        bg = _C["alt"] if i % 2 else _C["white"]
        for j, key in enumerate(("ticker", "real_from", "base_from", "base"), start=1):
            c = ws.cell(row=row, column=j, value=r[key])
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _font(9, bold=(j == 1))
            c.alignment = _align("left" if j in (1, 4) else "center")
        row += 1
    note = ws.cell(row=row + 1, column=1, value=(
        "Recent period uses REAL fund returns from 'Real from'; earlier history is the proxy "
        "base (geo/asset index × exposure, leverage financed at ^IRX+0.5%, net of TER). Proxies "
        "are total-return, converted to the reporting currency; modeled, not an exact replication. "
        "Factor ETFs additionally carry a factor tilt (Developed FF SMB/HML/RMW/MOM loadings, shown in the "
        "base column) so their value/momentum/quality tilt is preserved pre-inception."))
    note.font = _font(8, italic=True, color=_C["muted"])
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)


def _testfol_sheet(wb, testfol, byinst=None) -> None:
    """Sheet with the testfol.io lines (ticker + allocation %) to paste per
    portfolio. Proxies/leverage syntax follow the KB (SIM tickers, ?L=, CASHX
    for the leverage funding). For each portfolio we show:
      1. a per-instrument breakdown (each fund → the testfol code(s) it maps to,
         so one instrument may expand into several codes, e.g. an efficient-core
         fund → equity + bond − CASHX);
      2. the aggregated lines (codes summed across instruments) + a copy-paste
         one-liner ready to paste into the site."""
    byinst = byinst or {}
    ws = wb.create_sheet("Testfolio")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["green"]
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 10

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="TESTFOL.IO — proxies & weights per portfolio")
    t.font = _font(14, bold=True, color=_C["white"])
    t.fill = _fill(_C["header"])
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value=(
        "USD proxies (SIM = long backfilled history). Efficient-core funds (NTSG/NTSZ/NTSX) "
        "expand into equity + bond − CASHX (leverage funding); 2x funds (CL2/LVWC) use ?L=. "
        "TER baked as ?E=. One instrument may map to several codes.")
        ).font = _font(8, italic=True, color=_C["muted"])
    ws.merge_cells("A3:E3")
    ws.cell(row=3, column=1, value=(
        "Managed futures (MFEH/DBMFE): blend 50% DBMFSIM + 50% KMLMSIM (the two trend SIMs "
        "diverge a lot — half-and-half cuts model risk; both ~1992+). "
        "Carry (UEQC/CRRY): NO testfol ticker exists (carry ≠ trend); DBMFSIM shown is only a "
        "rough stand-in — the correct method is a custom CSV: CRRY = BNP index BNPIF73P + CASHX − 0.66% "
        "(0.34 TER + 0.32 swap), UEQC = 2.5x BCOMF3T − 2.5x BCOM + CASHX then UBS index from 2015.")
        ).font = _font(8, italic=True, color=_C["amber"])
    ws.merge_cells("A4:E4")
    ws.cell(row=4, column=1, value=(
        "Inception limiters: factor SIMs MTUM/QUAL/VLUE start ~2013 → factor-tilt portfolios "
        "limited to ~2013 on testfol. Leverage cost: CL2 (USA 2x) ≈ TER; LVWC (World 2x) real cost "
        "≈1.6% once div drag is doubled (per Rational Reminder). Always keep CASHX/SP financing.")
        ).font = _font(8, italic=True, color=_C["amber"])

    row = 5
    for name, lines in testfol.items():
        row = _section_header(ws, row, name.replace("_", " "), 5)

        # --- 1. Per-instrument breakdown -----------------------------------
        inst_rows = byinst.get(name)
        if inst_rows:
            for c, txt in ((1, "Ticker"), (2, "Instrument"), (3, "Weight"),
                           (4, "testfol code(s)"), (5, "Code wt")):
                h = ws.cell(row=row, column=c, value=txt)
                h.font = _font(9, bold=True, color=_C["white"])
                h.fill = _fill(_C["header"])
                h.border = _border()
            row += 1
            for i, ir in enumerate(inst_rows):
                bg = _C["alt"] if i % 2 else _C["white"]
                codes = ir["codes"] or [("—", 0.0)]
                first = row
                for j, (expr, cw) in enumerate(codes):
                    if j == 0:
                        tk = ws.cell(row=row, column=1, value=ir["ticker"])
                        tk.font = _font(9, bold=True)
                        tk.fill = _fill(bg); tk.border = _border()
                        nm = ws.cell(row=row, column=2, value=ir["name"])
                        nm.font = _font(9); nm.fill = _fill(bg); nm.border = _border()
                        wc = ws.cell(row=row, column=3, value=ir["weight"] / 100.0)
                        wc.number_format = "0.0%"; wc.alignment = _align("center")
                        wc.font = _font(9, bold=True); wc.fill = _fill(bg)
                        wc.border = _border()
                    else:
                        for cc in (1, 2, 3):
                            fillc = ws.cell(row=row, column=cc)
                            fillc.fill = _fill(bg); fillc.border = _border()
                    ec = ws.cell(row=row, column=4, value=expr)
                    ec.font = _font(9, color=_C["green"]); ec.fill = _fill(bg)
                    ec.border = _border()
                    cwc = ws.cell(row=row, column=5, value=cw / 100.0)
                    cwc.number_format = "0.0%"; cwc.alignment = _align("center")
                    cwc.font = _font(9); cwc.fill = _fill(bg); cwc.border = _border()
                    row += 1
                if len(codes) > 1:
                    ws.merge_cells(start_row=first, start_column=1,
                                   end_row=row - 1, end_column=1)
                    ws.merge_cells(start_row=first, start_column=2,
                                   end_row=row - 1, end_column=2)
                    ws.merge_cells(start_row=first, start_column=3,
                                   end_row=row - 1, end_column=3)
            row += 1

        # --- 2. Aggregated lines (paste block) -----------------------------
        sub = ws.cell(row=row, column=1, value="Aggregated (paste-ready)")
        sub.font = _font(9, bold=True, italic=True, color=_C["muted"])
        row += 1
        h1 = ws.cell(row=row, column=1, value="testfol code")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        h2 = ws.cell(row=row, column=5, value="Alloc")
        for c in (h1, h2):
            c.font = _font(9, bold=True, color=_C["white"])
            c.fill = _fill(_C["header"])
            c.border = _border()
        row += 1
        total = 0.0
        for i, (tk, w) in enumerate(lines):
            bg = _C["alt"] if i % 2 else _C["white"]
            a = ws.cell(row=row, column=1, value=tk)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            a.font = _font(9, bold=True)
            a.fill = _fill(bg)
            a.border = _border()
            b = ws.cell(row=row, column=5, value=w / 100.0)
            b.number_format = "0.0%"
            b.alignment = _align("center")
            b.fill = _fill(bg)
            b.border = _border()
            total += w
            row += 1
        tot = ws.cell(row=row, column=1, value="Total")
        tot.font = _font(9, bold=True, color=_C["muted"])
        tc = ws.cell(row=row, column=5, value=total / 100.0)
        tc.number_format = "0.0%"
        tc.alignment = _align("center")
        tc.font = _font(9, bold=True, color=_C["muted"])
        row += 1
        # Copy-paste one-liner.
        oneliner = "  ".join(f"{tk} {w:g}" for tk, w in lines)
        lbl = ws.cell(row=row, column=1, value="Paste string")
        lbl.font = _font(8, bold=True, color=_C["muted"])
        cp = ws.cell(row=row, column=2, value=oneliner)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        cp.font = _font(8, color=_C["muted"])
        cp.alignment = _align("left")
        row += 2